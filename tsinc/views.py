from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from django.core import serializers
from django.http import HttpResponse
from django.db.models import Max
from .models import *
from .forms import *
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, user_passes_test
import pandas as pd
import openpyxl
from django.http import JsonResponse
import json
from .utils import print_data, print_calc_supervirsor, sort_list_point, modify_point_file, print_offer, print_notes, print_remission, print_order, generate_offer_
import os, shutil
from django.conf import settings
from django.contrib import messages
from django.db.models import Q, F
import math
from django.utils import timezone
from django.db.models import Case, When, Value, IntegerField, F
from django.views.decorators.http import require_POST
from django.utils.encoding import iri_to_uri  # Para codificar caracteres especiales
from collections import Counter
from django.contrib.auth.decorators import permission_required
from .utils import send_notification
from django.views.decorators.csrf import csrf_exempt

# Create your views here.

def save_activity(item,action,project,usersession):
    pass




def staff_required(user):
    return user.is_staff

def superuser_required(user):
    return user.is_superuser

def paginator(request, obj, number):
    paginator = Paginator(obj, number)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return page_obj


def verify_and_create(array, project):
        for item in array:
            i= GeneratedOffer.objects.filter(measure__icontains=item,project= project).first()
            if not i:
                GeneratedOffer.objects.create(measure=item,project=project)

@login_required
def home(request):

    is_staff = staff_required(request.user) # Verifica si el usuario es administrador

    if is_staff:
        return redirect("/show_all_offers/0")

    
    projects = Project.objects.filter(usersession=request.user)
    tabs = Tabs.objects.filter(project__in=projects)
    pages = Dasboard.objects.filter(tab__in=tabs)

    page_obj = paginator(request,projects,5)
    user = request.user  # El usuario actualmente autenticado
    session_key = request.session.session_key 
                       
    return render(request, 'home.html',{'page_obj': page_obj, 'username':user.username, 'session_key':session_key, 'tabs':tabs, 'pages':pages})

@permission_required('tsinc.delete_project', login_url='/accessdenied/')
@login_required
def delete_project(request, id):
    referer = request.META.get('HTTP_REFERER')
    object = get_object_or_404(Project, id=id)
    object.delete()
    return redirect(referer)

def increase_offer_code():# incrementa el consecutivo de la oferta
    code = OfferCode.objects.filter(name = "offer").first()
    
    if code:
        code_splited = code.code.split("-")

        new_code = int(code_splited[-1]) + 1

        code.code = f"{code_splited[0]}-{code_splited[1]}-{new_code}"

        code.save()
    else: # si el consecutivo de la oferta no existe entonce lo crea
        code = OfferCode.objects.create(name = "offer", code="C-1-100")
        code_splited = code.code.split("-")

        new_code = int(code_splited[-1]) + 1

        code.code = f"{code_splited[0]}-{code_splited[1]}-{new_code}"

        code.save()

    return code


@permission_required('tsinc.add_project', login_url='/accessdenied/')
@login_required
def create_project(request):
    referer = request.META.get('HTTP_REFERER')
    if request.method == 'GET':
       return render(request, 'createproject.html',{'form': CreateProject()})
    else:

        offer_code = increase_offer_code()

        # asesor = User.objects.filter(id = request.POST['asesor']).first()

        project = Project.objects.create(name = request.POST['name'], 
                                         code = offer_code.code,
                                         company_name = request.POST['company_name'],
                                         nit=request.POST['nit'], 
                                         asesor = request.POST['asesor'], 
                                         usersession=request.user )
        
        messages.success(request,"El proyecto se ha creado correctamente")

        array = [
            "SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",
            "DESCUENTO",
            "SUBTOTAL MENOS (-) DESCUENTO",
            "ADMINISTRACIÓN",
            "IMPREVISTROS",
            "UTILIDADES",
            "SUBTOTAL COSTO DIRECTO + INDIRECTO",
            "IVA/UTILIDAD",
            "VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE"
        ]

        verify_and_create(array, project)
        
        def create_folders(offers_folder):

            company_folder = Folder.objects.filter(name__icontains = project.company_name, parent = offers_folder ).first() # traesr la carpeta company
            
            if company_folder:# si la carpeta company exite
                Folder.objects.create(name=project.name, parent = company_folder, project = project) # crea la carpeta con el nombre del proyecto dentro de la carpeta company
            else:# si no exite 
                company_folder = Folder.objects.create(name = project.company_name, parent = offers_folder ) # crea la carpeta company
                Folder.objects.create(name=project.name, parent = company_folder, project = project) # crea la carpeta del projecto debajo de company

        offers_folder = Folder.objects.filter(name__icontains ="ofertas").first() # traer la carpeta ofertas

        if offers_folder:# si la carpeta ofertas existe
            create_folders(offers_folder)
     
        else: # si la carpeta ofertas no existe 
            offers_folder = Folder.objects.create(name="OFERTAS") # crea la carpeta "ofertas"
            create_folders(offers_folder) # crea las carpetas correpondiente

        return redirect("/")

def replace_cero(value):
    if value == 0:
        value = 1
    return value

def save_stat_prod(products):
    # products = Product.objects.all()
    product_stats = []
    for product in products:

        info_dict = {}
        
        orderproduct = OrderProduct.objects.filter(product=product).all()
        total_entrys = OrderEntry.objects.filter(product__in= orderproduct).count()
        total_shippeds = ProductSent.objects.filter(product = product).count() 
             
        if total_entrys < total_shippeds:
            rotations = total_entrys
        elif total_shippeds < total_entrys:
            rotations = total_shippeds
        else:
            rotations = total_shippeds

        if product.quantity > product.min_stock:

            out_stock = 1

        else:

            out_stock = 0

        rotations = rotations

        info_dict['product'] = product
        info_dict['out_stock'] = out_stock
        info_dict['rotations'] = rotations

        product_stats.append(info_dict)


    return product_stats
    
        # print(f"producto:{product.model}--Rotaciones:{rotations}--Fuera stock:{out_stock}")
        # prod_stat = ProductStatictics.objects.filter(product = product).first()
        # if prod_stat:
        #     prod_stat.rotations = rotations
        #     prod_stat.out_stock = out_stock
        #     prod_stat.save()
        # else:   
        #     ProductStatictics.objects.create(product=product, rotations = rotations, out_stock = out_stock)


   

@permission_required('tsinc.view_product', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def show_inventory(request):

    search = request.GET.get('search')
    currency = Trm.objects.filter(currency = "usd").first()
    
    if search:
        page_obj = Product.objects.filter(Q(product_name__icontains= search) | Q(model__icontains= search) | Q(code__icontains= search))
    else:
        page_obj = Product.objects.all()
        # page_obj = Product.objects.filter(product_name__icontains= search)

    product_files = File.objects.filter(product__isnull = False)
    products = Product.objects.all()

    product_files_info = [
    {'product_id': product.id, 
     'quantity': sum(1 for file in product_files if file.product.id == product.id)
    } 
    for product in products
    ]   

    is_staff = staff_required(request.user)

    page_obj = paginator(request,page_obj,10)
    return render(request, 'product.html',{'page_obj': page_obj,
                                           'is_staff':is_staff,
                                           'currency':currency,
                                           'product_files_info':product_files_info
                                           })


@login_required
def dashboard(request, id):
    dashboard = Dasboard.objects.get(id=id)
    tab = dashboard.tab
    project = tab.project
    tabs = project.tabs.all()
    pages = Dasboard.objects.filter(tab__in=tabs).select_related('tab')
    panelitems = PanelItems.objects.all()
    # categorys = Category.objects.all()
    # subcategorys = Subcategory.objects.all()
    items = Items.objects.all()
    labels = Labels.objects.all()
    categorys = Category.objects.filter(parent__isnull = True)
    subcategorys = Category.objects.filter(parent__isnull = False)
   
    return render(request, 'dashboard.html',{'tab':tab, 
                                             'panelitems':panelitems, 
                                             'items': items,
                                             'dashboard':dashboard,
                                             'labels':labels, 
                                             'categorys':categorys, 
                                             'subcategorys':subcategorys, 
                                             'pages':pages, 
                                             'form':SearchForm()})

@login_required
def product_search(request):
    if request.method == 'POST':
        raw_data = request.body
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        results = Product.objects.filter(product_name__icontains= data['search']).values()
        results = list(results)
        return JsonResponse(results,safe=False)

@user_passes_test(staff_required,login_url='/accessdenied/')   
@login_required
def create_product(request):
    if request.method == 'GET':
       return render(request, 'createproduct.html',{'form': CreateProduct()})
    else:
        print(request.POST)
        if request.POST['iva'] == 'on':
            iva = True
        else:
            iva = False

        project = Product.objects.create(code = request.POST['code'], 
                                         product_name = request.POST['product_name'], 
                                         factory_ref= request.POST['ref'],
                                         brand= request.POST['brand'], 
                                         model= request.POST['model'], 
                                         location= request.POST['location'],
                                         quantity= request.POST['quantity'],
                                         purcharse_price= request.POST['purcharse_price'],
                                         sale_price= request.POST['sale_price'],
                                         min_stock= request.POST['min_stock'],
                                         point= request.POST['point'],
                                         observation= request.POST['observation'],
                                         description= request.POST['description'],
                                         iva=iva)
        return redirect('/product')


def verify_code_point_des(description):
    code = "NA"
    point = "NA"
    descrip = "NA"
    min_stock = 0
    
    if pd.notna(description):
        description_list = description.split(",")
        if len(description_list) == 4:
            code = description_list[0]
            min_stock = description_list[1]
            point = description_list[2]
            descrip =  description_list[3]
    
    return code,min_stock,point,descrip
        
@user_passes_test(staff_required,login_url='/accessdenied/')    
@login_required   
def upload_products(request):
    if request.method == 'POST':
        form = UploadProducts(request.POST, request.FILES)
    
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                original = request.POST['original']
            except:
                original = None

            df = pd.read_excel(excel_file, engine='openpyxl')
            try:
                if not original:
            
                    for _,row in df.iterrows():
                        if  pd.notna(row['Nombre del Producto / Servicio (obligatorio)']) and pd.notna( row['Referencia de Fábrica']) and pd.notna(row['Modelo'] ) and pd.notna(row['Marca'] ):
                            product_name = row['Nombre del Producto / Servicio (obligatorio)'] 
                            factory_ref = row['Referencia de Fábrica'] 
                            model = row['Modelo'] 
                            brand = row['Marca'] 
                            location = row['UBICACIÓN'] if pd.notna(row['UBICACIÓN']) else 'NA'
                            quantity = row['CANTIDAD'] if pd.notna(row['CANTIDAD']) else 0
                            description = row['ORBSERVACION']
                            sale_price = row['Precio de venta 12'] if pd.notna(row['Precio de venta 12']) else 0
                            purcharse_price = row['Precio de venta 11'] if pd.notna(row['Precio de venta 11']) else 0
                            
                            product_exist = Product.objects.filter(model = model, brand = brand).exists()
                            
                            code,min_stock,point,descrip = verify_code_point_des(description)
                                
                            if product_exist:
                                product= Product.objects.filter(model=model, brand=brand).first()
                                product.product_name = product_name
                                product.factory_ref = factory_ref
                                product.model = model
                                product.brand = brand
                                product.location = location
                                product.quantity = quantity
                                product.sale_price = sale_price
                                product.purcharse_price = purcharse_price
                                product.code = code
                                product.point = point
                                product.description = descrip
                                product.min_stock = min_stock
                                product.save()
                            
                            else:
                                                            
                                product = Product.objects.create(
                                product_name=product_name,
                                factory_ref=factory_ref,
                                model=model,
                                brand=brand,
                                location=location,
                                quantity= quantity,
                                sale_price = sale_price,
                                purcharse_price = purcharse_price,
                                code = code,
                                point = point,
                                min_stock = min_stock,
                                description=descrip,
                                )
                    messages.success(request, 'Productos subidos y creados exitosamente.')
                else:

                    for _,row in df.iterrows():
                        code = row['code'] if pd.notna(row['code']) else 'NA'
                        product_name = row['product_name']
                        factory_ref = row['factory_ref']
                        model = row['model']
                        sale_price = row['sale_price']
                        purcharse_price = row['purcharse_price']
                        brand = row['brand']
                        location = row['location'] if pd.notna(row['location']) else 'NA'
                        quantity = row['quantity']
                        min_stock = row['min_stock']
                        point = row['point'] if pd.notna(row['point']) else 'NA'
                        observation = row['observation'] if pd.notna(row['observation']) else 'NA'
                        description = row['description'] if pd.notna(row['description']) else 'NA'
                        iva = row['iva']

                        product_exist = Product.objects.filter(model = model, brand = brand ).exists()

                        if product_exist:
                            product= Product.objects.get(model=model, brand=brand)
                            product.code = code
                            product.product_name = product_name
                            product.factory_ref = factory_ref
                            product.model = model
                            product.brand = brand
                            product.location = location
                            product.quantity = quantity
                            product.sale_price = sale_price
                            product.purcharse_price = purcharse_price
                            product.point = point
                            product.min_stock = min_stock
                            product.observation = observation
                            product.description = description
                            product.iva = iva
                            product.save()
                        else:
                            product = Product.objects.create(
                                    code = code,
                                    product_name = product_name,
                                    factory_ref = factory_ref,
                                    model = model,
                                    brand = brand,
                                    location = location,
                                    quantity = quantity,
                                    sale_price = sale_price,
                                    purcharse_price = purcharse_price,
                                    min_stock = min_stock,
                                    description = description,
                                    point = point,
                                    iva = iva,
                            )


                    messages.success(request, 'Productos subidos y creados exitosamente.')
            except:
                messages.error(request, 'Algo ha salido mal asegurese de que la tabla este contruida correctamente, no tenga parametro vacios, y marcar la casilla original si la tabla es la original de la base de datos.')


                
            # messages.error(request, 'No se pudieron subir los productos a la base de datos, porfavor verifica que tu tabla tenga todos los parametros esten llenos')

            
            return redirect('/uploadproducts')
    else:
        form = UploadProducts()
    return render(request, 'uploadproducts.html', {'form': form})

@login_required
def download_products(request):
    # Consulta todos los datos del modelo
    queryset = Product.objects.all().values()
    # Convierte el queryset a un DataFrame de pandas
    df = pd.DataFrame(queryset)

    # Crear una respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tabla.xlsx'

    # Guardar el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')
    return response



@login_required
def tabs(request, id):
    if request.method == 'GET':
        project = Project.objects.get(id=id)
        tabs = project.tabs.all()
        
        # tabs = offer.tabs.all()
        return render(request,'tabs.html',{'tabs': tabs, 
                                           'form': CreateTab(), 
                                           'project':project
                                           })
    else:

        referer = request.META.get("HTTP_REFERER")
        
        project = Project.objects.get(id=id)
        chest_type = True
        if request.POST['chest_type'] == "2":
            chest_type = False 

        tab_name = request.POST['tab_name']

        tab = Tabs.objects.filter(tab_name = tab_name, project = project).first()

        
        if tab:  
            messages.error(request,f"El tablero con el nombre de {tab.tab_name} ya existe, elige otro nombre")
        
        else:
            tab = Tabs.objects.create(tab_name = request.POST['tab_name'], 
                                  chest_type = chest_type, 
                                  controller= request.POST['controller'], 
                                  project= project)
    
        return redirect(referer)

@login_required
def delete_tab(request, id):
    object = get_object_or_404(Tabs, id=id)
    project = object.project
    object.delete()
    return redirect(f"/project/tabs/{project.id}/")

@login_required
def download_points(request,id):
    project =Project.objects.get(id=id)
    tabs = project.tabs.all()
    pages = Dasboard.objects.filter(tab__in=tabs).select_related('tab')
    items = Items.objects.filter(dashboard__in=pages).select_related('dashboard') 
    labels = Labels.objects.filter(dashboard__in=pages).select_related('dashboard')  
    referer = request.META.get("HTTP_REFERER")
    
    parent_labels = [ label for label in labels if label.relationship != 'None'] 
        
    data = [ {
            'tab_name': label.dashboard.tab.tab_name,
            'unit_name': label.value,
            'related_items': [ item for item in items if label.relationship == item.relationship and item.img.product != None]
        }
        for label in parent_labels
        ]
    

    if not data:
        messages.error(request,f"¡Ha ocurrido un error al momento de generar el archivo!, asegurese de haber creado equipos en la dashboard")
        return redirect(referer)
    
    
    # print(pd.DataFrame(data))
    
    
    # sheet.cell(row=i+3, column=9, value=controller.point) 

    def create_sheet(tabs):
        sheets = [ workbook.create_sheet(title=tab.tab_name) for tab in tabs if tab.chest_type]
        return sheets
    try:
        workbook = openpyxl.Workbook()
        # # sheet = workbook.active
        # # sheet.title = 'TAB01'

        sheet_to_delete = workbook.active
        workbook.remove(sheet_to_delete)
    
   
    
        sheets = create_sheet(tabs)
        
        
        sheet = Sheet.objects.filter(project = project).all()
        
        sheet.delete()
    
        c_quantity = 0

        for i in range(0,len(tabs)):
            if tabs[i].chest_type:
                c_quantity += 1
                print_data(data,tabs[i],sheets[i],project)

        
        try: 
            sv_tab = [tab for tab in tabs if not tab.chest_type][0]
        except:
            sv_tab = None
        
        if sv_tab:
            sheet_supervisor = workbook.create_sheet(title='Supervisor')
            print_calc_supervirsor(sheet_supervisor,sv_tab,c_quantity,project)
        
        path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'points','Points.xlsx')
        workbook.save(path)

        
        date = datetime.datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=Points_{project.code}_{date.strftime("%d-%m-%Y")}.xlsx"

        workbook.save(response)
        return response
    except TypeError as e:
        messages.error(request,f"¡Ha ocurrido un error al momento de generar el archivo! {e}")
        return redirect(referer)
    

    
@login_required
def save_items(request):

    if request.method == 'GET':
        return redirect('/')
    else:
        raw_data = request.body
        # print(f"desde save_item: {raw_data}"  )
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        for value in data['values']:
            img_obj = PanelItems.objects.get(id=value['pk'])
            try:
                item_exist = Items.objects.get(id_code = value['id_code'])
            except ObjectDoesNotExist:
                item_exist = None
            
            if item_exist:
                item_exist.x = value['x']
                item_exist.y = value['y']
                item_exist.zindex = int(value['zindex'])
                item_exist.relationship=value['relationship']
                item_exist.tag = value['tag']
                # item_exist.width = float(value['width'].replace("px",""))          
                item_exist.save()
            else:
                dashboard = get_object_or_404(Dasboard, id=int(data['dashboard_id'])) 
                new_item = Items(id_code=value['id_code'],
                                 x=value['x'],y=value['y'],
                                #  width =float(value['width'].replace("px","")),
                                 relationship=value['relationship'], 
                                 img=img_obj, tag=value['tag'], dashboard= dashboard)
                new_item.save()

        for label in data['labels']:
            try:
                label_exist = Labels.objects.get(id_code = label['id_code'])
            except ObjectDoesNotExist:
                label_exist = None

            if label_exist:
                label_exist.value = label['value']
                label_exist.x = label['x']
                label_exist.y = label['y']
                label_exist.zindex = int(label['zindex'])
                label_exist.width = float(label['width'].replace("px",""))
                label_exist.height = float(label['height'].replace("px",""))
                label_exist.relationship=label['relationship']
                label_exist.save()
            else:
                dashboard = get_object_or_404(Dasboard, id=int(data['dashboard_id']))                         
                new_label = Labels(id_code=label['id_code'],
                                   value=label['value'],
                                   x=label['x'],y=label['y'],
                                   width =float(label['width'].replace("px","")),
                                   height=float(label['height'].replace("px","")), 
                                   relationship=label['relationship'],
                                   dashboard= dashboard )
                new_label.save()
       
        return JsonResponse({'mensaje': 'Datos guardados con éxito!'})
    
@login_required
def create_page(request,id):
    referer = request.META.get("HTTP_REFERER")
    if request.method == 'GET':
        tab = get_object_or_404(Tabs,id=id)
        pages_tab = tab.dashboards.all()
        # page = Dasboard.objects.all()
        paginator = Paginator(pages_tab, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'createpage.html',{'page_obj': page_obj, 'form':CreatePage(), 'tab':tab})
    else:
        tab = Tabs.objects.get(id=id)
        Dasboard.objects.create(name = request.POST['name'], tab=tab)
        return redirect(referer)

@login_required
def delete_page(request, id):
    referer = request.META.get("HTTP_REFERER")
    object = get_object_or_404(Dasboard, id=id)
    tab = object.tab
    object.delete()
    return redirect(referer)
@login_required
def delete_item(request):
    if request.method == 'POST':
        raw_data = request.body
        # print(f"desde delete_item: {raw_data}")
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        for item in data['id_code']:
            try: 
                item_exist = Items.objects.get(id_code = item)     
            except ObjectDoesNotExist:
                item_exist = None
            if item_exist:
                item_exist.delete()
            else:
                labels = get_object_or_404(Labels,id_code=item)
                labels.delete()
        return JsonResponse({'mensaje': 'eliminado con éxito!'})
    

@login_required
def total(request):
    total_points_list = []
    dashboard = Dasboard.objects.get(id=request.GET['id'])
    items = Items.objects.filter(dashboard = dashboard).all()
   
    dashboard_products = [ item.img.product for item in items if item.img.product != None]

    total = list(set(dashboard_products))
 
    total_products = []

    points = ['BI','BO','AI','AO']

    subtotal = 0
    
    def clean_point(point):
        point_cleaned = ""
        for i in point:
            if not i.isnumeric():
                point_cleaned += i
        return point_cleaned

    for p in total:
        count = 0
        for product in dashboard_products:
            if p.id == product.id:
                count += 1
        total_product = {}
        total_product['product']  = p
        total_product['pointtype'] = clean_point(p.point)
        total_product['quantity'] = count
        total_product['total_price'] = p.sale_price * count
        subtotal += p.sale_price * count
        total_products.append(total_product)
    
    for point in points:
        total = 0
        for product in total_products:
            if product['pointtype'] == point:
                total += product['quantity']
        total_points={}
        total_points['pointtype'] = point
        total_points['quantity'] = total
        total_points_list.append(total_points)

    return render(request, 'total.html', {'dashboard':dashboard,'items':items, 'total_products':total_products, 'subtotal':subtotal, 'total_points_list':total_points_list})



def handle_uploaded_file(file, path):
    with open(path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

@permission_required('tsinc.add_panelitems', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def upload_svg(request):

    if request.method == 'POST' and request.FILES.getlist('files'):#verifica si el metodo es post y si exiten archivos
        # tag = request.POST.get('tag') #etiqueta 
        category_id = request.POST.get('category') #etiqueta 
        files = request.FILES.getlist('files') #archivos
        category = Category.objects.filter(id=category_id).first()
        folder_name = category.name

        def validation_panelitem(file):#valida que exita el item de panel
            panelitem_exists = PanelItems.objects.filter(name=file.name).exists()
            return panelitem_exists
            
       
        def validation_folder(folder_name):#valida que exista el folter
            exist = Folders.objects.filter(name= folder_name).exists()
            return exist
            
            
        panelitem_exist = list(map(validation_panelitem, files)) #con la funcion map valida cada uno de los archivos y devuelve una lista de booleanos

    
        if not any(panelitem_exist):
            
            exist = validation_folder(folder_name)

            if not exist:# si no existe el folder entra y lo crea
                folder = Folders.objects.create(
                        name= folder_name,
                        path = f'items/{folder_name}'
                    )
            else:# De lo contrario lo busca en la base de datos

                folder = Folders.objects.get(name=folder_name)
                 
            def save_panelitem(file):
                
                # product = get_object_or_404(Product,model=file.name.split(".")[0])
                product = Product.objects.filter(model=file.name.split(".")[0]).first()# busca en la tabla de prodcutos, por el nombre de producto en el svg

                if product:
                    panelitem = PanelItems(name=file.name,  
                                        img=f'items/{folder_name}/{file.name}', 
                                        category_id=category_id, product=product, folder= folder)
                    panelitem.save()
                else:
                    panelitem = PanelItems(name=file.name,  
                                        img=f'items/{folder_name}/{file.name}', 
                                        category_id=category_id,folder= folder)
                    panelitem.save()
                    
                return panelitem
                    
            saved_panelitem = list(map(save_panelitem,files))# guarda informacion de los archivos en la base de datos

            panelitem_names = [ item.name for item in saved_panelitem] #lista de elementos guardados
            
            def filter_no_porduct(panelitem):# función para filtrar lo que no tienen un producto asignado
                if not panelitem.product:
                    return panelitem

            panelitem_no_product = list(filter(filter_no_porduct,saved_panelitem))
            panelitem_no_product_str = [ panelitem.name for panelitem in panelitem_no_product]
                
        # print(os.path.join(settings.BASE_DIR, 'tsinc','static', 'img', folder_name))        
            for file in files:
                # Define la ruta donde quieres guardar los archivos
                upload_dir = os.path.join(settings.MEDIA_ROOT, 'items', folder_name)
                os.makedirs(upload_dir, exist_ok=True)  # Crea la carpeta si no existe
                file_path = os.path.join(upload_dir, file.name)
                handle_uploaded_file(file, file_path)
           
           
            messages.success(request, f"{', '.join(panelitem_names)} Elementos subidos exitosamente.")

        

            if len(panelitem_no_product) != 0:
                messages.info(request, f"{', '.join(panelitem_no_product_str)} Elementos no han sido relacionado con ningun producto.")
        else:

            panelitems = [ file.name for file in files if PanelItems.objects.filter(name= file.name).exists()]

            messages.error(request, f"Los elementos: {', '.join(panelitems)} ya existen en la base de datos") 
          
              
               
        return redirect('/uploadsvg')

    folders = Folders.objects.all()
    categories = Category.objects.filter( parent__isnull = False )
    return render(request, 'uploadsvg.html', {'folders':folders,'categories':categories})


@permission_required('tsinc.view_folders', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def files_folders(request):
     
    panelitems = PanelItems.objects.all()
    folders = Folders.objects.all()
    
    return render(request, 'filesfolders.html', {
        'folders':folders,
        'panelitems': panelitems
    } )

@permission_required('tsinc.delete_panelitem', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def delete_file_item(request, id):

    try:
        panelitem = PanelItems.objects.get(id=id)
    except ObjectDoesNotExist:
        panelitem = None
    
    file_path = os.path.join(settings.MEDIA_ROOT, os.path.normpath(panelitem.img) )

    if panelitem and os.path.exists(file_path):
        panelitem.delete()
        os.remove(file_path)
        messages.success(request,f"El acrchivo {panelitem.name} ha sido eliminado correctamente")
    else:
        messages.error(request,'El archivo no se encontró!!')
        
    

    return redirect('/filesfolders')


@permission_required('tsinc.delete_folders', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def delete_folder_item(request, id):
    
    try:
        folder = Folders.objects.get(id=id)
    except ObjectDoesNotExist:
        folder = None
    
    folder_path = os.path.join(settings.MEDIA_ROOT,os.path.normpath(folder.path))

  
    if folder and os.path.exists(folder_path):
        folder.delete()
        shutil.rmtree(folder_path)
        messages.success(request,f"La carpeta {folder.name} ha sido eliminada correctamente")
    else:
        messages.error(request,'La carpeta no se encontró!!')
        
    return redirect('/filesfolders')
@login_required
def edit_tab(request):
    id = int(request.GET.get('id'))
    tab = get_object_or_404(Tabs,id=id)
    if request.method == 'POST':
        tab_name = request.POST.get('tab')
        option = request.POST.get('option')
        if option == '1':
            controller = "LG BECON CONTROLLER"
        elif option == '2':
            controller = "JCI FACILITY EXPLORER"
        elif option == '3':
            controller = "JCI METASYS"
        tab.tab_name = tab_name
        tab.controller = controller
        tab.project = tab.project
        tab.save()
        messages.success(request,f"El tablero {tab.tab_name} ha sido guardado correctamente.")
        return redirect(f'/edittab/?id={tab.id}')
    
    return render(request,'edittab.html',{'tab':tab})

@login_required
def modify_points_file(request,id):

    referer = request.META.get("HTTP_REFERER")
    try:
        if request.method == "GET":
            project = Project.objects.get(id=id)
            sheets = project.sheet.all()
            sheet_id = request.GET.get("sheet")
            licenses = License.objects.filter(sheet__in =sheets)
            sv = False
            if sheet_id:
                sheet = Sheet.objects.filter(id=sheet_id).first()
                points = Points.objects.filter(sheet = sheet).all()
                if sheet.name == "Supervisor":
                    sv=True 
            else:
                sheet = sheets[0]
                points = Points.objects.filter(sheet = sheet).all()
            
            return render(request,"modifypoint.html",{'sheets':sheets, 
                                                    'points':points, 
                                                    'addcontroller':AddController(), 
                                                    'addlicense':AddLicense(), 
                                                    'sheet_id':sheet.id,
                                                    'project':project,
                                                    'licenses':licenses,
                                                    'sv':sv
                                                    }
                                                    
                                                    )
        else:
            def is_controller(string):
                for l in string:
                    if l == "C":
                        return True 
                return False
                
            sheet = Sheet.objects.filter(id=request.POST.get("sheet_id")).first()
            try:
                request_license = request.POST["license"]
            except:
                request_license = None
            if not request_license:
                controller = Product.objects.filter(id=request.POST["controller"]).first()
                
                if not  controller.point == None:
                    points = controller.point.split("-")
                    sort_points = sort_list_point(points)

                    Points.objects.create(name = controller.model,
                                        is_controller = is_controller(controller.code),
                                        eu=sort_points[0],
                                        ed=sort_points[1],
                                        sa=sort_points[2],
                                        sd=sort_points[3],
                                        sc=sort_points[4],
                                        sheet = sheet
                                        )
                else:
                    Points.objects.create(name = controller.model,
                                        is_controller = is_controller(controller.code),
                                        sheet = sheet
                                        )

            else:
                # print(request.POST["license"])
                license = Product.objects.filter(id=request.POST["license"]).first()
                License.objects.create(name=license.model,description=license.description, sheet=sheet)

            return redirect(referer)
    except:
        messages.error(request,"¡Ha ocurrido un error para generar la información!, asegurese de haber descargado el archivo de puntos previamente.")
        return redirect(referer)


def delete_controller(request, id):
    controller = Points.objects.filter(id=id).first()
    sheet = controller.sheet
    controller.delete()
    return redirect(f"/modifypoints/{sheet.project.id}?sheet={sheet.id}")

def download_point_excel(request,id):
    project = Project.objects.get(id=id)
    path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'points','Points.xlsx')
    workbook = openpyxl.load_workbook(path)
    
    for sheet_name in workbook.sheetnames:
        modify_point_file(project,workbook[sheet_name])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Points.xlsx'

    workbook.save(response)
    
    return response

def modify_supervisor(request, id):
    return render(request,"modifysuper.html")

def delete_license(request, id):
    license = License.objects.filter(id=id).first()
    sheet = license.sheet
    license.delete()
    return redirect(f"/modifypoints/{sheet.project.id}?sheet={sheet.id}")

@login_required
def download_offer(request,id):

    try:
        project =Project.objects.get(id=id)
        workbook = openpyxl.Workbook()
        # # sheet = workbook.active
        # # sheet.title = 'TAB01'

        sheet_to_delete = workbook.active
        workbook.remove(sheet_to_delete)

        sheet = workbook.create_sheet(title="Oferta")
        sheet_notes = workbook.create_sheet(title="NOTAS Y ACLARACIONES")

        print_offer(sheet,project)
        print_notes(sheet_notes,project)
        # path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'points','Points.xlsx')
        # workbook.save(path)

        path = os.path.join(settings.MEDIA_ROOT, 'offers', f"{project.id}")

        # Crea la carpeta si no existe
        os.makedirs(path, exist_ok=True)

        file_path = os.path.join(path,f"offer_{project.code}.xlsx")
        
        workbook.save(file_path)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Oferta_{project.code}.xlsx'
        
        workbook.save(response)

        return response
    except TypeError as e:
        messages.error(request,f"¡Ha ocurrido un error al momento de generar el archivo!{e}, asegurese de haber creado tableros, paginas y equipos en la dashboard. si el problema persiste contacte a la empresa.")
        return redirect('/')


@login_required
def generate_offer(request,id):
    referer = request.META.get("HTTP_REFERER")
    project =Project.objects.get(id=id)
    tabs = Tabs.objects.filter(project = project).exists()

    
                
    if not tabs:
        messages.info(request,"La oferta no tiene tableros de control creados!")
       
        array = [
            "SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",
            "DESCUENTO",
            "SUBTOTAL MENOS (-) DESCUENTO",
            "ADMINISTRACIÓN",
            "IMPREVISTROS",
            "UTILIDADES",
            "SUBTOTAL COSTO DIRECTO + INDIRECTO",
            "IVA/UTILIDAD",
            "VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE"
        ]

        verify_and_create(array)

    
        return redirect(referer)
    try:
        generate_offer_(project)
        messages.success(request,"La oferta ha sido generada correctamente")
    except TypeError as e :
        messages.error(request,f"Algo ha salido mal asegurese de haber generados los puntos de control, error:{e}")
    
    return redirect(referer)


@login_required
def edit_page(request):
    id = int(request.GET.get('id'))
    page = get_object_or_404(Dasboard,id=id)
    if request.method == 'POST':
        page_name = request.POST.get('page_name')
        page.name = page_name
        page.save()
        messages.success(request,f"Cambios realizados correctamente a {page.name}")
        return redirect(f'/editpage/?id={page.id}')
    
    return render(request,'editpage.html',{'page':page})

@permission_required('tsinc.change_project', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def edit_project(request, id):
    project = get_object_or_404(Project,id=id)
    if request.method == 'POST':
        project_name = request.POST.get('project_name')
        company_name = request.POST.get('company_name')
        nit = request.POST.get('nit')
        project.name = project_name
        project.company_name = company_name
        project.nit = nit
        project.save()
        messages.success(request,f"Cambios realizados correctamente")
        return redirect(f'/editproject/?id={project.id}')
    
    return render(request,'editproject.html',{'project':project})


@login_required
def delete_product(request, id):
    product = Product.objects.filter(id=id).first()
    product.delete()
    return redirect(f"/product/")

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def edit_product(request,id):
    product = Product.objects.filter(id = id).first()
    if request.method == 'POST':
        product = Product.objects.filter(id = id).first()
        code = request.POST.get('code')
        product_name = request.POST.get('product_name')
        factory_ref = request.POST.get('factory_ref')
        model = request.POST.get('model')
        sale_price = request.POST.get('sale_price')
        purcharse_price = request.POST.get('purcharse_price')
        brand = request.POST.get('brand')
        location = request.POST.get('location')
        quantity = request.POST.get('quantity')
        point = request.POST.get('point')
        observation = request.POST.get('observation')
        description = request.POST.get('description')
        measure = request.POST.get('measure')
        min_stock = request.POST.get('min_stock')
        product.code = code
        product.product_name = product_name
        product.factory_ref = factory_ref
        product.model = model
        product.sale_price = sale_price
        product.purcharse_price = purcharse_price
        product.brand = brand
        product.location = location
        product.quantity = quantity
        product.point = point
        product.observation = observation
        product.description = description
        product.min_stock = min_stock
        product.measure = measure
        product.save()
        messages.success(request,f"Cambios realizados correctamente")
        return redirect(f'/editproduct/{product.id}')
    
    is_staff = staff_required(request.user)
    return render(request,'editproduct.html',{'product':product, 'is_staff':is_staff})
    


@login_required
def add_product_to_box(request,id): 
    referer = request.META.get("HTTP_REFERER")
    product = Product.objects.filter(id=id).first()
    user = request.user
    if ProductBox.objects.filter(product = product, usersession = request.user).exists():
        productbox = ProductBox.objects.filter(product = product).first()
        productbox.quantity +=1
        productbox.save() 
        messages.success(request,f"El producto {product.model} ya se encontraba en el carrito ahora tienes {productbox.quantity}.")
    else:
        ProductBox.objects.create(product = product, quantity = 1, price =product.sale_price, usersession = user) 
        messages.success(request,f"El producto {product.model} ha sido agregado al carrito.")
    return redirect(referer)



@login_required
def add_product(request,id):
    if request.method == "POST":
        print(request.POST.get(f"quantity{id}"))
    # print(request.GET.get(f"quantity{id}"))
    # productbox = ProductBox.objects.filter(id = id).first()
    # productbox.quantity +=1
    # productbox.save() 
        # messages.success(request,f"El producto {product.model} ya se encontraba en el carrito ahora tienes {productbox.quantity}.")
    return redirect('/createremission/')

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def subtract_product(request,id): 
    referer = request.META.get('HTTP_REFERER')
    productbox = ProductBox.objects.filter(id = id).first()
    productbox.delete()
        # messages.success(request,f"El producto {product.model} ya se encontraba en el carrito ahora tienes {productbox.quantity}.")
    return redirect(referer)


def increase_code():
    remission_code = OfferCode.objects.filter(name = "remission").first()
    if remission_code:
        new_code = int(remission_code.code) + 1

        remission_code.code = new_code

        remission_code.save()
    else:
        remission_code = OfferCode.objects.create(name= "remission", code ="1" )
        
        new_code = int(remission_code.code) + 1

        remission_code.code = new_code

        remission_code.save()


    return remission_code

@permission_required('tsinc.add_remission', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def create_remission(request,project_id = None): 
   
    project = Project.objects.filter(id=project_id).first()
    projects = Project.objects.all()

    if request.method == 'GET':
        path = f"createremission"

        productbox = ProductBox.objects.filter(usersession = request.user).all()
        return render(request,'createremission.html',{'productbox':productbox,
                                                      'form':CreateRemission,
                                                      'path':path,
                                                      'project_from_offer':project,
                                                      'projects':projects,
                                                      })
    else:
        productbox = ProductBox.objects.filter(usersession = request.user).all()

        
        for productb in productbox:
            if productb.quantity > productb.product.quantity:
                messages.error(request,f"No hay sufiente cantidad del producto {productb.product.model} en el inventario")
                return redirect(f"/createremission/{project.id}")

        remission_code = increase_code()

        remission = Remission.objects.create(city = request.POST['city'],
                                            number = remission_code.code,
                                            order_number = request.POST['order_number'],
                                            company = request.POST['company'],
                                            nit = request.POST['nit'],
                                            location = request.POST['location'],
                                            contruction_site = request.POST['contruction_site'],
                                            supplier = request.POST['supplier'],
                                            responsible = request.POST['responsible'],
                                            observation = request.POST['observation'],
                                            project_id = request.POST.get('project_id'),
                                            usersession = request.user
                                            )
        products = []
        for productb in productbox:
            productsent = ProductSent.objects.create(
                product = productb.product,
                quantity = productb.quantity,
                price = productb.price,
                tab = productb.tab,
                section = productb.section,
                remission = remission
            )
            product = Product.objects.filter(id= productsent.product.id).first()
            products.append(product)
            product.quantity -= productsent.quantity
            product.save()
        
        # seve_stat_prod(products)

        messages.success(request,f"La remisión ha sido generada correctamente.")

        return redirect(f"/show_remission/{remission.id}")
    
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def clean_productbox(request):
    referer = request.META.get('HTTP_REFERER')  
    productbox = ProductBox.objects.filter(usersession = request.user).all()
    productbox.delete()
    return redirect(referer)

@permission_required('tsinc.view_remission', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def remissions(request):  
    remissions = Remission.objects.filter().order_by('-date').all()
    search = request.GET.get('search')
    if search:
        remissions = Remission.objects.filter(Q(company__icontains= search) | Q(nit__icontains= search) | Q(number__icontains= search) | Q(project__name__icontains= search))
    
    remission_files = File.objects.filter(remission__isnull = False)
    
    remission_files_info = [
    {'remission_id': remission.id, 
     'quantity': sum(1 for file in remission_files if file.remission.id == remission.id)
    } 
    for remission in remissions
    ]         

    remissions = paginator(request,remissions,10)

    return render(request,'remissions.html',{'remissions':remissions, 
                                             'remission_files_info':remission_files_info
                                             })
@permission_required('tsinc.delete_remission', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_remission(request,id):  
    remission = Remission.objects.filter(id = id).first()
    remission_prod = ProductSent.objects.filter(remission = remission ).all()
    referer = request.META.get('HTTP_REFERER')
    for prod in remission_prod:
        product = Product.objects.filter(id = prod.product.id).first()
        product.quantity += prod.quantity
        product.save()
    remission.delete()
    return redirect(referer)

@permission_required('tsinc.view_remission', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def show_remission(request,id):  
    remission = Remission.objects.filter(id=id).first()
    products = ProductSent.objects.filter(remission= remission).all()
    remissionfiles = File.objects.filter(remission = remission).all()
    folder = Folder.objects.filter(project = remission.project).first()

    projects = Project.objects.all()
    from_show_remission = True
    return render(request,'show_remission.html',{'products':products, 
                                                   'remission':remission, 
                                                   'remissionfiles':remissionfiles,
                                                   'from_show_remission':from_show_remission,
                                                   'projects':projects,
                                                   'folder':folder
                                                   })

@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def product_shipped(request):  
    products_shipped = ProductSent.objects.all()[:10]
    search = request.GET.get('search')
    if search:
        products = Product.objects.filter(Q(product_name__icontains= search) | Q(model__icontains= search))
        products_shipped = ProductSent.objects.filter(product__in = products)
    is_staff = staff_required(request.user)   
    return render(request,'productshipped.html',{'products':products_shipped,'is_staff':is_staff})

def calc_t_quantity_price(order):# calcula la cantidad total y el precio total de la orden de compra
    
    orderproducts = OrderProduct.objects.filter(order = order).all()
    t_quantity = 0
    t_price = 0

    for product in orderproducts:
        t_quantity += product.quantity 
        t_price += round(product.price * product.quantity,1)

    order.total_quantity = t_quantity
    order.total_price = t_price

    order.save()

def calc_progress(order):

    order_entrys = OrderEntry.objects.filter(order = order).all()
    arrived = 0
   
    for entry in order_entrys:
        arrived += entry.quantity
    
    if order.total_quantity > 0:
        progress = math.ceil((arrived * 100) / order.total_quantity)
    else:
        progress = 0

    order.progress = progress
    order.save()

def increase_code_order():



    code = OfferCode.objects.filter(name = "order").first()

    if code:

        code_splited = code.code.split("-")

        new_code = int(code_splited[-1]) + 1

        code.code = f"{code_splited[0]}-{new_code}"

    else:

        code = OfferCode.objects.create(name = "order", code="OC-1" )

        code_splited = code.code.split("-")

        new_code = int(code_splited[-1]) + 1

        code.code = f"{code_splited[0]}-{new_code}"

    code.save()

    return code
@permission_required('tsinc.add_purcharseorder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_order(request, project_id = None):


    project = Project.objects.filter(id = project_id).first()

    projects = Project.objects.all()

    if request.method == 'GET':

        path = "createorder"
        productbox = ProductBox.objects.filter(usersession = request.user).all()
        
        return render(request,"createorder.html",{'productbox':productbox,
                                                  'form':CreateOrder(),
                                                   'path':path,
                                                   'project_from_offer':project,
                                                   'projects':projects
                                                     })

    else:
        productbox = ProductBox.objects.filter(usersession = request.user).all()
        tracking = request.POST['tracking']
        supplier = request.POST['supplier']
        nit = request.POST['nit']
        address = request.POST['address']
        phone = request.POST['phone']
        customer = request.POST['customer']
        cost_center = request.POST['cost_center']
        inspector = request.POST['inspector']
        supervisor = request.POST['supervisor']
        observation = request.POST['observation']
        currency = request.POST['currency']
        project = request.POST.get('project_id')
    
        code = increase_code_order()

        order = PurcharseOrder.objects.create(code = code.code,
                                              tracking = tracking,
                                              supplier = supplier,
                                              nit = nit,
                                              address = address,
                                              phone = phone,
                                              customer = customer,
                                              cost_center = cost_center,
                                              inspector = inspector,
                                              supervisor = supervisor,
                                              observation = observation,
                                              currency = currency,
                                              usersession = request.user,
                                              project_id= project
                                             )
        
        for product in productbox:

            OrderProduct.objects.create(
                product = product.product,
                quantity = product.quantity,
                price = product.price,
                order = order
            )
        
        calc_t_quantity_price(order)

        messages.success(request,f"La orden de compra ha sido generada correctamente.")
            
        return redirect(f"/orderproductinfo/{order.id}")
    
@login_required
def save_car(request):
    if request.method == "POST":
        raw_data = request.body
        # print(f"desde save_item: {raw_data}"  )
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        for product in data['values']:
            productbox = ProductBox.objects.filter(id= product['id']).first()
            productbox.quantity = product['quantity']
            productbox.price = product['price']
            productbox.save()
        return JsonResponse({'mensaje': 'Datos guardados con éxito!'})


@permission_required('tsinc.view_purcharseorder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def purcharse_order(request):  
    orders = PurcharseOrder.objects.filter().order_by('progress').all()
    search = request.GET.get('search')
    if search:
        orders = PurcharseOrder.objects.filter(Q(code__icontains= search) | Q(supplier__icontains= search) | Q(tracking__icontains= search) | Q(project__name__icontains= search))

    orders = paginator(request,orders,10)

    order_files = File.objects.filter(order__isnull = False)

    order_files_info = [ # calcula la cantidad de archivos que pertenecen a la orden
    {'order_id': order.id, 
     'quantity': sum(1 for file in order_files if file.order.id == order.id)
    } 
    for order in orders
    ] 

    order_invoices = OrderInvoice.objects.all()
    
    billing = [ #calcula el pregreso en porcentaje de facturación
    {
        'order_id': order.id,
        'progress': round(
            sum(invoice.value_paid for invoice in order_invoices if order.id == invoice.order.id) * 100 / replace_cero(order.total_price),
            1
        )
    }
    for order in orders
    ]

    
    return render(request,'purcharseorders.html',{'orders':orders, 
                                                  'order_files_info':order_files_info,
                                                  'billing':billing
                                                  })

@user_passes_test(staff_required,login_url='/accessdenied/')
@login_required
def purcharse_order_products(request):  
    purcharse_order_products = OrderProduct.objects.all()[:10]
    search = request.GET.get('search')
    if search:
        products = Product.objects.filter(Q(product_name__icontains= search) | Q(model__icontains= search))
        purcharse_order_products = OrderProduct.objects.filter(product__in = products)  

    is_staff = staff_required(request.user)   
    return render(request,'purcharseorderproducts.html',{'products':purcharse_order_products,'is_staff':is_staff})


def calc_info_order_product(orderproducts,orderentry):
    info_per_product = []
    for product in orderproducts:
        t_quantity = 0
        t_price = 0
        p={}
        for entry in orderentry:
            if product.id == entry.product.id:
                t_quantity += entry.quantity

        p['id'] = product.id
        p['t_quantity'] = t_quantity
        p['leftover'] = product.quantity - t_quantity      
        info_per_product.append(p)
    return info_per_product

@permission_required('tsinc.view_purcharseorder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')   
@login_required
def order_product_info(request,id):  
    order = PurcharseOrder.objects.filter(id=id).first()
    products = OrderProduct.objects.filter(order= order).all()
    orderentry = OrderEntry.objects.filter(order = order).all()
    folder = Folder.objects.filter(project = order.project).first()

    orderfiles = File.objects.filter(order = order).all()
    projects = Project.objects.all()
    
    most_recent_entry = OrderEntry.objects.filter(order = order).order_by("-date").first()
    
    intial_date = order.date

    if most_recent_entry:
        final_date = most_recent_entry.date
        
        # intial_date = datetime.date(2023, 6, 1)
        # final_date = datetime.datetime(2023, 7, 25)
        difference = final_date - intial_date
        delivery_time = difference.days
    else:
        delivery_time = 0

    arrived = 0
    d_price = 0

    for entry in orderentry:
        arrived += entry.quantity 
        d_price += entry.product.price * entry.quantity
    
    
    
    total_info = {
        't_price':round(order.total_price,2),
        'd_price':round(d_price,2),
        'r_price':round(order.total_price - d_price,2),
        't_quantity':order.total_quantity,
        'arrived':arrived,
        'leftover':round(order.total_quantity - arrived,2)
    }

    badge = ""

    if order.currency:
        badge = "COP"
    else:
        badge = "USD"
    
    info_per_product = calc_info_order_product(products,orderentry)
    
    is_staff = staff_required(request.user)

    from_show_order = True

    return render(request,'show_order.html',{'products':products, 
                                                   'order':order, 
                                                   'orderentry':orderentry,
                                                   'total_info':total_info,
                                                   'info_per_product':info_per_product,
                                                   'delivery_time':delivery_time,
                                                   'badge':badge,
                                                   'orderfiles':orderfiles,
                                                   'is_staff':is_staff,
                                                   'from_show_order':from_show_order,
                                                   'projects':projects,
                                                   'folder':folder
                                                   
                                                   })


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def download_remission(request,id):

    try:
        remission =Remission.objects.get(id=id)
        workbook = openpyxl.Workbook()
        # # sheet = workbook.active
        # # sheet.title = 'TAB01'

        sheet_to_delete = workbook.active
        workbook.remove(sheet_to_delete)

        sheet = workbook.create_sheet(title="Remision")

        print_remission(sheet,remission)
        # path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'points','Points.xlsx')
        # workbook.save(path)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=Remision_{remission.number}.xlsx"

        workbook.save(response)

        return response
    except TypeError as e:
        messages.error(request,f"¡Ha ocurrido un error al momento de generar el archivo!.{e}")
        return redirect('/remissions/')

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def download_order(request,id):

    try:
        order =PurcharseOrder.objects.get(id=id)
        workbook = openpyxl.Workbook()
        # # sheet = workbook.active
        # # sheet.title = 'TAB01'

        sheet_to_delete = workbook.active
        workbook.remove(sheet_to_delete)

        sheet = workbook.create_sheet(title="Orden de compra")

        print_order(sheet,order)
        # path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'points','Points.xlsx')
        # workbook.save(path)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=Orden_de_compra_{order.code}.xlsx"

        workbook.save(response)

        return response
    except TypeError as e:
        messages.error(request,f"¡Ha ocurrido un error al momento de generar el archivo!.{e}")
        return redirect('/purcharseorder/')

@permission_required('tsinc.change_remission', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def edit_remission(request,id):
    referer = request.META.get('HTTP_REFERER')
    remission = get_object_or_404(Remission,id=id)
    products = ProductSent.objects.filter(remission = remission).all()
    if request.method == 'POST':
        city = request.POST.get('city')
        company = request.POST.get('company')
        nit = request.POST.get('nit')
        location = request.POST.get('location')
        contruction_site = request.POST.get('contruction_site')
        supplier = request.POST.get('supplier')
        project = request.POST.get('project')
        responsible = request.POST.get('responsible')
        observation = request.POST.get('observation')
        project_id = request.POST.get('project_id')
        remission.city = city
        remission.company = company
        remission.nit = nit
        remission.contruction_site = contruction_site
        remission.supplier = supplier
        remission.location = location
        remission.project = project
        remission.responsible = responsible
        remission.observation = observation
        remission.project_id = project_id
        remission.save()
        messages.success(request,f"Cambios realizados correctamente")
        return redirect(referer)
    
    return render(request,'edit_remission.html',{'remission':remission, 'products':products})

@permission_required('tsinc.change_purcharseorder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')     
@login_required
def edit_order(request,id):
    referer = request.META.get('HTTP_REFERER')
    order = get_object_or_404(PurcharseOrder,id=id)
    orderproducts = OrderProduct.objects.filter(order = order).all()
    if request.method == 'POST':
        tracking = request.POST.get('tracking')
        supplier = request.POST.get('supplier')
        nit = request.POST.get('nit')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        customer = request.POST.get('customer')
        cost_center = request.POST.get('cost_center')
        inspector = request.POST.get('inspector')
        supervisor = request.POST.get('supervisor')
        observation = request.POST.get('observation')
        currency = request.POST.get('currency')
        project_id = request.POST.get('project_id')
        order.tracking = tracking
        order.supplier = supplier
        order.nit = nit 
        order.address = address
        order.phone = phone
        order.customer = customer
        order.cost_center = cost_center
        order.inspector = inspector
        order.supervisor = supervisor
        order.observation = observation
        order.currency = currency
        order.project_id = project_id
        order.save()
        messages.success(request,f"Cambios realizados correctamente")
        return redirect(referer)
    

    return render(request,'editorder.html',{'order':order,'orderproducts':orderproducts})

      
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_order_entry(request,id):

    order = get_object_or_404(PurcharseOrder,id=id)
    orderproducts = OrderProduct.objects.filter(order = order).all()
    orderentry = OrderEntry.objects.filter(order=order).order_by("-date").all()

    info_per_product = calc_info_order_product(orderproducts,orderentry)

    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        quantity = request.POST.get('quantity')
        tracking = request.POST.get('tracking')
        if not product_id == 'None':
            orderproduct = OrderProduct.objects.filter(id=product_id).first()
            # if orderproduct.order.trm != 0:
            #     price_dollar = round(orderproduct.price / orderproduct.order.trm,2)
            #     orderentry = OrderEntry.objects.create(order = order, 
            #                           product = orderproduct, 
            #                           quantity=quantity,
            #                           price = price_dollar,
            #                           tracking=tracking,
            #                           )
            # else:
            orderentry = OrderEntry.objects.create(order = order, 
                                    product = orderproduct, 
                                    quantity=quantity,
                                    price = orderproduct.price,
                                    tracking=tracking,
                                    )
            
            entries = OrderEntry.objects.filter(product__product__id = orderproduct.product.id).all()
            
            currency = Trm.objects.filter(currency = "usd").first()
            prices = []
            for entry in entries:
                if entry.order.currency:
                    if currency.value > 0:
                        price_usd = round(entry.price / currency.value,2)
                        prices.append(price_usd)
                else:
                    prices.append(entry.price)
            
            highest_price =  max(prices)
            product = Product.objects.filter(id = orderproduct.product.id).first()
            product.purcharse_price = highest_price
            product.quantity += int(orderentry.quantity)
            product.save()
            calc_progress(order)
            # seve_stat_prod([product])
            messages.success(request,f"La entrada ha sido creada correctamente")

        else:

            messages.error(request,f"No puedes generar mas entradas, Todos los productos han sido entregados!")
        return redirect(f'/createorderentry/{id}')
    is_staff = staff_required(request.user)
    return render(request,'createorderentry.html',{'order':order,
                                                   'orderproducts':orderproducts, 
                                                   'orderentry':orderentry,
                                                   'info_per_product':info_per_product,
                                                   'is_staff':is_staff
                                                   })


    
@permission_required('tsinc.delete_purcharseorder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_order(request,id):
    referer = request.META.get("HTTP_REFERER")
    order = PurcharseOrder.objects.filter(id = id).first()
    order.delete()
    # save_activity(f"ORDEN {order.code}","DELETE",order.project,request.user)
    return redirect(referer)


@login_required
def delete_order_product(request,id):  
    referer = request.META.get("HTTP_REFERER")
    orderproduct = OrderProduct.objects.filter(id = id).first()
    order = PurcharseOrder.objects.filter(id = orderproduct.order.id).first()
    orderproducts = OrderProduct.objects.filter(order = order).all()
    orderproduct.delete()
    calc_t_quantity_price(order)
    calc_progress(order)
    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_order_entry(request,id):  
    orderentry = OrderEntry.objects.filter(id = id).first()
    order = PurcharseOrder.objects.filter(id = orderentry.order.id).first()
    product = Product.objects.filter(id = orderentry.product.product.id).first()
    product.quantity -= orderentry.quantity
    product.save()
    orderentry.delete()
    calc_t_quantity_price(order)
    calc_progress(order)
    return redirect(f"/createorderentry/{order.id}")


@login_required
def entry_info(request,id): 
    entry = OrderEntry.objects.filter(id=id).first()
    return render(request,"entryinfo.html",{'entry':entry})



def get_total_info_order(purcharseorders, entrys):
    currency = Trm.objects.filter(currency="usd").first()

    total_invoiced_ = OrderInvoice.objects.filter(order__in = purcharseorders)

    total_price = 0
    total_delivered = 0
    total_invoiced_usd = 0
    total_invoiced_cop = 0

    for order in purcharseorders:
        if order.currency:
            if currency.value > 0:
                total_price += round(order.total_price / currency.value,2)
        else:
            total_price += order.total_price

        for entry in entrys:
            if order.id == entry.order.id:
                if order.currency:
                    total_delivered += (entry.price/currency.value) * entry.quantity
                else:
                    total_delivered += entry.price * entry.quantity

    for orderinvoice in total_invoiced_:
        if orderinvoice.order.currency:
            total_invoiced_usd += orderinvoice.value_paid/currency.value
        else:
            total_invoiced_usd += orderinvoice.value_paid

        

    total_info ={
        'total_orders': round(total_price,2),
        'total_delivered':round(total_delivered,2),
        'total_remaining': round(total_price - total_delivered,2),
        'total_invoiced': round(total_invoiced_usd,2),
        'total_remaining_invoiced': round(total_price - total_invoiced_usd,2),

    }

    return total_info


@user_passes_test(superuser_required,login_url='/accessdenied/')
@login_required
def order_statictics(request):

  
    most_recent_entry = OrderEntry.objects.all().order_by("-date")[:5]
    most_recent_order = PurcharseOrder.objects.all().order_by("-date")[:5]
    purcharseorders = PurcharseOrder.objects.exclude(progress=100)
    entrys = OrderEntry.objects.filter(order__in = purcharseorders)
  
    # difference = final_date - intial_date
    # delivery_time = difference.days

    total_info = get_total_info_order(purcharseorders,entrys)


    pending_orders = [{
        'order':order,
        'ago':order.date
    }
    for order in purcharseorders if order.progress != 100
    ][:5]

   
    
    return render(request,"orderstatictics.html",{
                                                  'recent_entrys':most_recent_entry,
                                                  'recent_orders':most_recent_order,
                                                  'total_info':total_info,
                                                  'pending_orders':pending_orders,
                                        
                                                    })

@login_required
def product_info(request,id): 
    product = Product.objects.filter(id=id).first()
    products = Product.objects.all()
    orderproduct = OrderProduct.objects.filter(product = product).all()
    orderentry = OrderEntry.objects.filter(product__in= orderproduct ).order_by("-date")[:5]
    products_shipped = ProductSent.objects.filter(product = product).order_by('-remission__date')[:5]
    total_entrys = OrderEntry.objects.filter(product__in= orderproduct).count()
    total_shippeds = ProductSent.objects.filter(product = product).count() 
    rotations = 0
    if total_entrys < total_shippeds:
        rotations = total_entrys
    elif total_shippeds < total_entrys:
        rotations = total_shippeds
    else:
        rotations = total_shippeds

    product_files = File.objects.filter(product = product).all()
 
    
    return render(request,"productinfo.html",{'product':product, 
                                              'orderentry':orderentry, 
                                              'products_shipped':products_shipped, 
                                              'total_entrys':total_entrys,
                                              'total_shippeds':total_shippeds,
                                              'rotations':rotations,
                                              'product_files':product_files,
                                              })

@login_required
def add_entry_inventory(request,id):
    entry = OrderEntry.objects.filter(id = id).first()
    product = Product.objects.filter(id = entry.product.product.id).first()
    if not entry.added:
        entry.added = True
        entry.save()
        product.quantity += entry.quantity
        product.save()
    return redirect(f"/createorderentry/{entry.order.id}")


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_remission_product(request,id):  
    remissionproduct = ProductSent.objects.filter(id = id).first()
    referer = request.META.get("HTTP_REFERER")
    product = Product.objects.filter(id = remissionproduct.product.id).first()

    product.quantity += remissionproduct.quantity

    product.save()
    
    remissionproduct.delete()

    return redirect(referer)

@user_passes_test(superuser_required,login_url='/accessdenied/')
@login_required
def product_statictics(request):
    products = Product.objects.all()
    prod_out_stock = ProductStatictics.objects.order_by("out_stock")[:10]
    prod_rotations = ProductStatictics.objects.order_by("-rotations")[:10]
    expensive_prod = Product.objects.order_by("-sale_price")[:10]
    cheaper_prod = Product.objects.order_by("sale_price")[:10]

    product_stats = save_stat_prod(products)

    product_out_stock_ = sorted(product_stats, key=lambda x: x['out_stock'])[:10]
    prod_rotations_ = sorted(product_stats, key=lambda x: x['rotations'], reverse=True)[:10]

    total_inventory = 0
    total_inventory_cop = 0

    for product in products:
        total_inventory += product.purcharse_price * product.quantity
        total_inventory_cop += product.purcharse_price_cop * product.quantity
    
    total_inventory = round(total_inventory,2)
    total_inventory_cop = round(total_inventory_cop,2)
    return render(request,"productstatictics.html",{
                                                    'total_inventory':total_inventory,
                                                    'total_inventory_cop':total_inventory_cop,
                                                    'exp_prod':expensive_prod,
                                                    'cheaper_prod':cheaper_prod,
                                                    'product_out_stock_':product_out_stock_,
                                                    'prod_rotations_':prod_rotations_
                                                    })

@user_passes_test(superuser_required,login_url='/accessdenied/')
@login_required
def remission_statictics(request):
    remissions = Remission.objects.order_by('-date')[:5]
    prod_remission = ProductSent.objects.order_by('-remission__date')[:5]

    return render(request,"remissionstatictics.html",{'remissions':remissions,'prod':prod_remission})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_car_remission(request,id):
    referer = request.META.get("HTTP_REFERER")
    productbox = ProductBox.objects.filter(usersession = request.user).all()
    remission = Remission.objects.filter(id = id).first()

    for productb in productbox:
            if productb.quantity > productb.product.quantity:
                messages.error(request,f"No hay sufiente cantidad del producto {productb.product.model} en el inventario")
                return redirect(referer)
    products = []
    for productb in productbox:
        productsent = ProductSent.objects.create(
            product = productb.product,
            quantity = productb.quantity,
            price = productb.price,
            remission = remission
        )
        product = Product.objects.filter(id= productsent.product.id).first()
        products.append(product)
        product.quantity -= productsent.quantity
        product.save()
    
    # seve_stat_prod(products)

    return redirect(referer)
    

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_car_order(request,id):
    referer = request.META.get("HTTP_REFERER")
    productbox = ProductBox.objects.filter(usersession = request.user).all()
    order = PurcharseOrder.objects.filter(id = id).first()


    for product in productbox:

            OrderProduct.objects.create(
                product = product.product,
                quantity = product.quantity,
                price = product.price,
                order = order
            )
        
    calc_t_quantity_price(order)
    calc_progress(order)

    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def duplicate_remission(request,id):
    remission = Remission.objects.filter(id = id).first()
    referer = request.META.get('HTTP_REFERER')
    try:
        remission_code = increase_code()

        Remission.objects.create(
            number = remission_code.code,
            city = remission.city,
            company = remission.company,
            nit = remission.nit,
            location = remission.location,
            project = remission.project,
            responsible = remission.responsible,
            observation = remission.observation,
            order_number = remission.order_number,
            usersession = request.user
        )
        messages.success(request,"La remisión se ha duplicado con éxito")
    except TypeError as e:
         messages.error(request,f"Se ha presentado el siguiente error: {e}")
    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required 
def duplicate_order(request,id):
    order = PurcharseOrder.objects.filter(id = id).first()
    
    try:
        code = increase_code_order()

        PurcharseOrder.objects.create(
                                    code = code.code,
                                    tracking = order.tracking,
                                    supplier = order.supplier,
                                    nit = order.nit,
                                    address = order.address,
                                    phone = order.phone,
                                    customer = order.customer,
                                    cost_center = order.cost_center,
                                    inspector = order.inspector,
                                    supervisor = order.supervisor,
                                    observation = order.observation,
                                    usersession = request.user  
                                        )
        messages.success(request,"La orden de compra se ha duplicado con éxito")
       
    except TypeError as e:
         
        messages.error(request,f"Se ha presentado el siguiente error: {e}")

    return redirect("/purcharseorder/")

def access_denied(request):
    return render(request,"accessdenied.html")

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def carpage(request):
    productbox = ProductBox.objects.filter(usersession = request.user).all()
    path = "carpage"
    return render(request,"carpage.html",{'productbox':productbox,'path':path })

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def save_trm(request):

    if request.method == "POST":
        if Trm.objects.filter(currency = "usd").exists():
            currency = Trm.objects.filter(currency = "usd").first()
            currency.value = request.POST.get('trmvalue')
            currency.save()
        else:
            Trm.objects.create(currency="usd",value = request.POST.get('trmvalue'))
        
        products = Product.objects.all()
        currency = Trm.objects.filter(currency = "usd").first()
    
        for product in products:
            sale_price_cop = round(currency.value * product.sale_price,2)
            purcharse_price_cop = round(currency.value * product.purcharse_price,2)
            product.sale_price_cop = sale_price_cop
            product.purcharse_price_cop = purcharse_price_cop
            product.save()

    messages.success(request,f"Se ha establecido el TRM de: {currency.value} a todos los productos")    

    return redirect("/product/")


def upload_file(file,file_without_space,path):
    upload_dir = os.path.join(settings.MEDIA_ROOT, path)
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file_without_space)
    handle_uploaded_file(file, file_path)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_remission_file(request,id):
    remission = Remission.objects.filter(id = id).first()
    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")

            remission_file = File.objects.filter(name=file_without_space,remission = remission).first()
            folder_name = "remission_files"

            if remission_file:

                upload_file(file,file_without_space,remission_file.path)
           
                messages.info(request,f"El archivo {remission_file.name} ya existe, ha sido reemplazado")
                return redirect(f"/uploadremissionfile/{remission.id}")
            else:
                remission_file = File.objects.create(name = file_without_space,
                                                            path=f"{folder_name}/{remission.number}",
                                                            remission=remission,
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,remission_file.path)
            
            messages.success(request,f"El archivo {remission_file.name} ha sido cargado correctamente")

        return redirect(f"/uploadremissionfile/{remission.id}")
    
    
    
    return render(request,"uploadremissionfile.html",{'form':UploadFile, 'remission':remission})

@permission_required('tsinc.view_file', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def view_file(request,id):
    file = get_object_or_404(File,id=id)
        # Construir la URL del archivo
    file_url = f"{settings.MEDIA_URL}{file.path}/{file.name}"

    # Codificar la URL para manejar caracteres especiales (como acentos)
    file_url = iri_to_uri(file_url)
    
    return render(request, "view_file.html", {"file_url": file_url, 
                                                  "file_name": file.name,
                                                  })
    
@permission_required('tsinc.view_delete', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_file(request,id):
    referer = request.META.get("HTTP_REFERER")
    file = get_object_or_404(File,id=id)
    file_path = os.path.join(settings.MEDIA_ROOT, file.path, file.name)
    
    if os.path.exists(file_path):
        os.remove(file_path) 
    
    file.delete()

    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_order_file(request,id):
    order = PurcharseOrder.objects.filter(id = id).first()
    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")
            order_file = File.objects.filter(name=file_without_space,order = order).first()
            folder_name = "order_files"
            if order_file:
                
                upload_file(file,file_without_space,order_file.path)

                messages.info(request,f"El archivo {order_file.name} ya existe, ha sido reemplazado")
                return redirect(f"/uploadorderfile/{order.id}")
            
            else:
                
                order_file = File.objects.create(name = file_without_space,
                                                        path = f"{folder_name}/{order.code}",
                                                        order=order,
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,order_file.path)

            messages.success(request,f"El archivo {order_file.name} ha sido cargado correctamente")

        return redirect(f"/uploadorderfile/{order.id}")
    
    return render(request,"uploadorderfile.html",{'form':UploadFile,  
                                                      'order':order})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_invoice_file(request,id):
    invoice = Invoice.objects.filter(id = id).first()
    referer = request.META.get("HTTP_REFERER")
    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")

            invoice_file = File.objects.filter(name=file_without_space,invoice = invoice).first()

            folder_name = "invoice_files"
            folder_id = f"{invoice.id}"

            if invoice_file:
                
                
                upload_file(file,file_without_space,invoice_file.path)

                messages.info(request,f"El archivo {invoice_file.name} ya existe, ha sido reemplazado")
                return redirect(referer)
            
            else:
                invoice_file = File.objects.create(name = file_without_space,
                                                            invoice=invoice,
                                                            path = f"{folder_name}/{folder_id}",
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,invoice_file.path)

            messages.success(request,f"El archivo {invoice_file.name} ha sido cargado correctamente")

        return redirect(referer)
    
    return render(request,"upload_invoice_file.html",{'form':UploadFile,  
                                                      'invoice':invoice})

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required   
def delete_order_file(request,id):
    order_file = get_object_or_404(File,id=id)
    referer = request.META.get("HTTP_REFERER")
    file_path = os.path.join(settings.MEDIA_ROOT, "order_files", f"{order_file.order.code}", order_file.name)
    
    if os.path.exists(file_path):
        os.remove(file_path) 
    
    order_file.delete()

    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_product_file(request,id):
    product = Product.objects.filter(id = id).first()
    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")
            
            product_file = File.objects.filter(name=file_without_space,product = product).first()
            folder_name = "product_files"
            if product_file:
                
                upload_file(file,file_without_space,product_file.path)


                messages.info(request,f"El archivo {product_file.name} ya existe, ha sido reemplazado")
                return redirect(f"/uploadproductfile/{product.id}")
            
            else:
                product_file = File.objects.create(name = file_without_space,
                                                            product=product,
                                                            path=f"{folder_name}/{product.id}",
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,product_file.path)

            messages.success(request,f"El archivo {product_file.name} ha sido cargado correctamente")

        return redirect(f"/uploadproductfile/{product.id}")
    
    is_staff = staff_required(request.user)
    
    return render(request,"uploadproductfile.html",{'form':UploadFile, 
                                                      'is_staff':is_staff, 
                                                      'product':product})

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required    
def delete_product_file(request,id):
    product_file = get_object_or_404(File,id=id)
    referer = request.META.get("HTTP_REFERER")
    file_path = os.path.join(settings.MEDIA_ROOT, "product_files", f"{product_file.product.id}", product_file.name)
    
    if os.path.exists(file_path):
        os.remove(file_path) 
    
    product_file.delete()

    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_invoice(request,id):

    project = Project.objects.filter(id=id).first()
    productbox = ProductBox.objects.filter(usersession = request.user).all()


    if request.method == "POST":
        number = request.POST.get('number')
        total_price = request.POST.get('total_price')
        iva = request.POST.get('iva')
        source_retention = request.POST.get('source_retention')
        ica_retention = request.POST.get('ica_retention')
        invoice = Invoice.objects.create(
            number = number,
            total_price = total_price,
            iva = iva,
            source_retention = source_retention,
            ica_retention = ica_retention,
            project = project,
            usersession = request.user
        )
        
        for productb in productbox:
            ProductInvoice.objects.create(
                product = productb.product,
                quantity = productb.quantity,
                price = productb.price,
                invoice = invoice
            )

        messages.success(request,"La factura ha sido creado correctamente")
        return redirect(f"/createinvoice/{project.id}")
    
    return render(request,"create_invoice.html",{
                                                  'project':project,
                                                  'productbox':productbox
                                                  })

@permission_required('tsinc.view_invoice', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def invoices(request):
    
    invoices = Invoice.objects.filter().order_by('-date').all()
    search = request.GET.get('search')
    if search:
        invoices = Invoice.objects.filter(Q(number__icontains= search) | Q(total_price__icontains= search) | Q(project__name__icontains= search))
    
    
    invoices = paginator(request,invoices,10)

    return render(request,"invoices.html",{
        'invoices':invoices
    })

def calc_total_invoice(invoices, order):
    total = 0
    ica_rete = 0
    iva = 0
    source_rete = 0
    total_invoice = {}
    for invoice in invoices:
        total += invoice.value_paid
        ica_rete = invoice.ica_retention
        source_rete = invoice.source_retention
        iva = invoice.iva

    total_invoice['total_value'] = total
    total_invoice['total_leftover'] = order.total_price - total
    total_invoice['ica_rete'] = ica_rete
    total_invoice['source_rete'] = source_rete
    total_invoice['iva'] = iva
     
    if total_invoice['total_leftover'] == 0:
        order.paid = True
        order.save()

    return total_invoice     
        
    
@permission_required('tsinc.add_orderinvoice', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_order_invoice(request,id):
    referer = request.META.get("HTTP_REFERER")
    order = get_object_or_404(PurcharseOrder,id=id)
    invoices = OrderInvoice.objects.filter(order = order).all()

    total_invoice = calc_total_invoice(invoices,order)

    order_invoice_files = File.objects.filter(order_invoice__in= invoices )

    if request.method == "POST":
        
        value_paid = request.POST.get('amount_to_pay')
        iva = request.POST.get('iva')
        source_retention = request.POST.get('source_retention')
        ica_retention = request.POST.get('ica_retention')
        if float(value_paid) <= order.total_price and float(value_paid) <= total_invoice['total_leftover']:
            OrderInvoice.objects.create(
                order = order,
                value_paid = value_paid,
                iva = iva,
                source_retention = source_retention,
                ica_retention = ica_retention,
                usersession = request.user
            )
        
            return redirect(referer)
        
        else:
            messages.error(request,f"El valor ingresado: {value_paid} es mayor del valor total de la orden: {order.total_price} o del valor restante {total_invoice['total_leftover']}")
            return redirect(referer)
      
    return render(request,"create_order_invoice.html",{
        'order':order,
        'invoices':invoices,
        'total_invoice':total_invoice,
        'order_invoice_files':order_invoice_files
    })


@permission_required('tsinc.delete_orderinvoice', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_order_invoice(request,id):
    referer = request.META.get("HTTP_REFERER")
    orderinvoice = get_object_or_404(OrderInvoice,id=id)
    orderinvoice.delete()
    return redirect(referer)


@permission_required('tsinc.view_project', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def show_all_offers(request,approved):

    if not approved:
        offers = Project.objects.filter(approved = False).all()
    
        search = request.GET.get('search')

        if search:
            offers = Project.objects.filter(Q(code__icontains= search) | Q(name__icontains= search) | Q(company_name__icontains= search), approved=False)
        
        approved = False
    else:
        
        offers = Project.objects.filter(approved = True).all()
    
        search = request.GET.get('search')

        if search:
            offers = Project.objects.filter(Q(code__icontains= search) | Q(name__icontains= search) | Q(company_name__icontains= search), approved=True)
        approved = True
    
    page_obj = paginator(request,offers,10)

  
                       
    return render(request,"show_all_offers.html",{'page_obj':page_obj, 'approved':approved})



def move_to_running_project(project):
    project_running = Folder.objects.filter(name__icontains = "PROYECTOS EN EJECUCION").first()

    if project_running:

        offers_folder = Folder.objects.filter(name__icontains = "OFERTAS").first()

        if offers_folder:

            company_folder_in_offers_folder = Folder.objects.filter(name__icontains= project.company_name,parent = offers_folder).first()

            if company_folder_in_offers_folder:

                project_folder = Folder.objects.filter(project = project, parent = company_folder_in_offers_folder).first()

                if project_folder:

                    company_folder_in_project_ronning = Folder.objects.filter(name__icontains = project.company_name,parent = project_running).first()

                    if company_folder_in_project_ronning:

                        project_folder.parent = company_folder_in_project_ronning
                        project_folder.save()
                    else:
                        new_company_folder_in_project_ronning = Folder.objects.create(name = project.company_name,parent = project_running)
                        project_folder.parent = new_company_folder_in_project_ronning
                        project_folder.save()
                else:
                    print("LA CARPETA RELACIONADA CON EL PROYECTO NO EXISTE!")

            else:
                print('CARPETA EMPRESA DENTRO DE OFERTA NO EXISTE!')
        else:

            print("OFERTAS NO EXISTE!")
    else:
        print("PROYECTOS EN EJECUCION NO EXISTE!")

            


def change_offer_status(request,id):
    project = Project.objects.filter(id=id).first()
    
    if not project.approved:
        project.approved = True
        
        move_to_running_project(project)

        project.save()
    else:
        project.approved = False
        project.save()

    return redirect("/show_all_offers/")




def folder_tree_func():

    root_folders =  Folder.objects.filter(parent__isnull=True)
    folders = Folder.objects.select_related('parent').order_by('parent_id', 'order')
   
    return folders

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def overview(request):
    

    folder_tree = folder_tree_func()
    
    context = {
        'folders': folder_tree, 
    }

    return render(request,"workspace.html",context)


def calc_offer_info_all_project(folders):

    
    folders_children = []
    for folder in folders:
        if folder.children:
            for folder_  in  folder.children.all():
                folders_children.append(folder_)     
        
    
    if folders_children:
        projects = [ folder.project  for folder in folders_children if folder.project]
    else:
        projects = [ folder.project  for folder in folders if folder.project]

    
    total_offer = 0
    total_invoiced = 0
    total_info_folder_customer = {}

    for project in projects:

        total_offer_ = GeneratedOffer.objects.filter(measure__icontains="VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",project = project).first()
        total_invoices_ = Invoice.objects.filter(project = project)
        if total_offer_:
            total_offer += total_offer_.total_value
        if total_invoices_:
            total_invoiced += sum(total_invoice.total_price for total_invoice in total_invoices_) 

    total_info_folder_customer['total_offer'] = total_offer
    total_info_folder_customer['total_invoiced'] = total_invoiced
    total_info_folder_customer['total_value_remaining'] = total_offer - total_invoiced

    return total_info_folder_customer
        
def calculate_and_save_percentage_of_tasks(project):

    tasks = Task.objects.filter(project = project).all()
    tasks_completed = Task.objects.filter(project = project, state = 'finalizado').all()

    if tasks:
        percentage_und_task = 100/len(tasks)
        porcentage_tasks = len(tasks_completed) * percentage_und_task


        project.progress = porcentage_tasks
        project.save()


def count_comments(tasks, comments, files):

    comments_per_task = [
        {
            'task':task,
            'number_of_comments':sum(1 for comment in comments if task.id == comment.task.id ),
            'number_of_files':sum(1 for file in files if task.id == file.task.id )

        }
        for task in tasks
    ]
    return comments_per_task

@permission_required('tsinc.view_folder', login_url='/accessdenied/')
@login_required
def overview_folder(request,id):
    folder = Folder.objects.filter(id = id).first()

    if not folder:
        return redirect("/overview/")


    projects = Project.objects.filter(folder__isnull=True)
    
    def sorted_files(files):

        files = [
            {
                'id':file.id,
                'name':file.name,
                'date':file.date,
                'user':file.usersession

            }
            for file in files
        ]
        files = sorted(files,key=lambda x:x['date'], reverse=True)[:10]

        return files 

    if folder.project:
        project = Project.objects.filter(id = folder.project.id).first()
        remissions = Remission.objects.filter(project  = project).all()
        orders = PurcharseOrder.objects.filter(project  = project).all()
        invoices = Invoice.objects.filter(project = project).all()
        remission_files = File.objects.filter(remission__project = project)[:5]
        invoice_file = File.objects.filter(invoice__project = project)[:5]
        folder_file = File.objects.filter(folder__project = project)[:5]
        generated_offer = GeneratedOffer.objects.filter(project= project).exists()
        pending_orders = PurcharseOrder.objects.filter(project  = project)
        entrys = OrderEntry.objects.filter(order__in = pending_orders)
        tasks = Task.objects.filter(project = project).all()
        comments = Comment.objects.filter(task__in = tasks)
        task_files = File.objects.filter(task__in = tasks)
    

        total_info_order = get_total_info_order(pending_orders,entrys) # calcular informacion resumen de ofertas y resumen de ordenes

        calculate_and_save_percentage_of_tasks(project)
        comments_per_task = count_comments(tasks, comments, task_files)

        files = sorted_files(remission_files)
        files += sorted_files(invoice_file)
        files += sorted_files(folder_file)
        offer_info = calc_offer_info(project)
        

    else:
        total_info_order = {}
        offer_info = {}
        project = None
        orders = None
        remissions = None
        invoices = None
        generated_offer = None
        tasks = None
        comments = None
        task_files = None
        comments_per_task = None
        folder_file = File.objects.filter(folder = folder).all()
        files = sorted_files(folder_file)

    if not folder.project:
        children = folder.children.all().order_by('date')[:5]
        children = folder.children.all()
        offer_info = calc_offer_info_all_project(children)

    else:
        children = None

    
    folder_tree = folder_tree_func()

    context = {

        'folders': folder_tree, 
        'folder': folder,
        'projects':projects,
        'remissions':remissions,
        'invoices':invoices,
        'orders':orders,
        'children':children,
        'files':files,
        'offer_info':offer_info,
        'generated_offer':generated_offer,
        'total_info_order':total_info_order,
        'tasks':tasks,
        'comments':comments,
        'comments_per_task':comments_per_task,
        'task_files':task_files
    }
    
    return render(request,'overview_folder.html',context)

@require_POST
def update_tree_order(request):
    try:
        data = json.loads(request.body)
        project_running_folder = Folder.objects.filter(name__icontains = "PROYECTOS EN EJECUCION").first()
        offers_folder = Folder.objects.filter(name__icontains = "OFERTAS").first()

        for item in data['order']:
            folder = Folder.objects.get(id=item['id'])
            folder.parent_id = item.get('parent', None)  # Actualiza el padre si existe
            if project_running_folder:
                if folder.project and folder.parent:
                    if folder.parent.parent:
                        if folder.parent.parent.id == project_running_folder.id:
                            project = Project.objects.filter(id = folder.project.id).first()
                            if not project.approved:
                                project.approved = True
                                project.save()
                        elif folder.parent.parent.id == offers_folder.id:
                            project = Project.objects.filter(id = folder.project.id).first()
                            if project.approved:
                                project.approved = False
                                project.save()
                     
            folder.order = item['order']
            folder.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@permission_required('tsinc.delete_folder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/')   
@login_required
def delete_folder(request,id):
    referer = request.META.get('HTTP_REFERER')
    folder = Folder.objects.filter(id = id).first()
    folder.delete()
    return redirect(referer)

@permission_required('tsinc.add_folder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def new_folder(request):
    referer = request.META.get('HTTP_REFERER')
    if request.method == "POST":
        name = request.POST.get("name")
        Folder.objects.create(name = name, color ="#ffffff" ,usersession = request.user , order = 0)
    return redirect(referer)

@permission_required('tsinc.change_folder', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def update_folder(request,id):
    referer = request.META.get('HTTP_REFERER')
    if request.method == "POST":
        name = request.POST.get("name")
        project_id = request.POST.get("project_id")
        color = request.POST.get("color")
        folder = Folder.objects.filter(id = id).first()     
        if not project_id == "none":
            folder.project_id = project_id
        else:
            folder.project = None
             
        folder.name = name
        folder.color = color
        folder.save()   
        
    return redirect(referer)

@permission_required('tsinc.view_invoice', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def show_invoice(request, id):
    invoice = Invoice.objects.filter(id=id).first()
    invoice_files = File.objects.filter(invoice= invoice).all()
    products_invoice = ProductInvoice.objects.filter(invoice = invoice).all()
    from_show_invoice = True
    return render(request,"show_invoice.html",{'invoice':invoice,
                                               'invoice_files':invoice_files,
                                               'products':products_invoice,
                                               'from_show_invoice':from_show_invoice
                                               })





@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_invoice(request,id):
    referer = request.META.get('HTTP_REFERER')
    invoice = Invoice.objects.filter(id = id).first()
    invoice.delete()
    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def update_invoice(request,id):
    invoice = Invoice.objects.filter(id = id).first()
    referer = request.META.get('HTTP_REFERER')
    if request.method == "POST":
        number = request.POST.get("number")
        total_price = request.POST.get("total_price")
        iva = request.POST.get("iva")
        source_retention = request.POST.get("source_retention")
        ica_retention = request.POST.get("ica_retention")

        invoice.number = number
        invoice.total_price = total_price
        invoice.iva = iva
        invoice.source_retention = source_retention
        invoice.ica_retention = ica_retention
        invoice.save()

        messages.success(request,"Se ha guardado correctamente")

        return redirect(referer)
    else:
        messages.error(request,"Algo ha salido mal y no se ha guardado")
    

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required   
def delete_invoice_file(request,id):
    invoice_file = get_object_or_404(File,id=id)
    referer = request.META.get("HTTP_REFERER")
    file_path = os.path.join(settings.MEDIA_ROOT, "invoice_files", f"{invoice_file.invoice.id}", invoice_file.name)
    
    if os.path.exists(file_path):
        os.remove(file_path) 
    
    invoice_file.delete()

    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def duplicate_invoice(request,id):
    invoice = Invoice.objects.filter(id = id).first()
    referer = request.META.get('HTTP_REFERER')
    try:
        Invoice.objects.create(
            number = invoice.number,
            total_price = invoice.total_price,
            iva = invoice.iva,
            source_retention = invoice.source_retention,
            ica_retention = invoice.ica_retention,
            usersession = request.user,
            project = invoice.project
        )
        messages.success(request,"La remisión se ha duplicado con éxito")
    except TypeError as e:
         messages.error(request,f"Se ha presentado el siguiente error: {e}")
    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_folder_file(request,id):
    folder = Folder.objects.filter(id = id).first()
    referer = request.META.get('HTTP_REFERER')

    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")

            folder_file = File.objects.filter(name=file_without_space,folder = folder).first()
            folder_name = "folder_files"

            if folder_file:

                upload_file(file,file_without_space,folder_file.path)
           
                messages.info(request,f"El archivo {folder_file.name} ya existe, ha sido reemplazado")
                return redirect(referer)
            else:
                folder_file = File.objects.create(name = file_without_space,
                                                            path=f"{folder_name}/{folder.name}",
                                                            folder=folder,
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,folder_file.path)
            
            messages.success(request,f"El archivo {folder_file.name} ha sido cargado correctamente")

        return redirect(referer)
    return render(request,"upload_folder_file.html",{'form':UploadFile,  
                                                      'folder':folder})

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def upload_order_invoice_file(request,id):
    order_invoice = OrderInvoice.objects.filter(id = id).first()
    referer = request.META.get('HTTP_REFERER')
    if request.method == "POST":
        form = UploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_without_space = file.name.replace(" ", "_")
            order_invoice_file = File.objects.filter(name=file_without_space, order_invoice = order_invoice).first()
            folder_name = "order_invoice_files"
            if order_invoice_file:
                
                upload_file(file,file_without_space,order_invoice_file.path)

                messages.info(request,f"El archivo {order_invoice_file.name} ya existe, ha sido reemplazado")
                return redirect(referer)
            
            else:
                order_invoice_file = File.objects.create(name = file_without_space,
                                                            order_invoice=order_invoice,
                                                            path=f"{folder_name}/{order_invoice.id}",
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,order_invoice_file.path)

            messages.success(request,f"El archivo {order_invoice_file.name} ha sido cargado correctamente")

        return redirect(referer)
    
    return render(request,"upload_order_invoice_file.html",{'form':UploadFile,  
                                                      'order_invoice':order_invoice})

@permission_required('tsinc.view_file', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def docs(request):
    
    docs = File.objects.filter().order_by('-date').all()
    search = request.GET.get('search')
    if search:
        docs = File.objects.filter(Q(name__icontains= search))
    
    
    docs = paginator(request,docs,10)

    return render(request,"docs.html",{
        'docs':docs
    })


def calc_offer_info(project):
    generated_offer = GeneratedOffer.objects.filter(project= project)
    remissions = Remission.objects.filter(project = project).all()    
    products_shipped = ProductSent.objects.filter(remission__in = remissions)
    invoices = Invoice.objects.filter(project= project)
    valor_total = GeneratedOffer.objects.filter(measure__icontains="VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",project = project).first()


    offer_info = {
        'total_product':sum(item.quantity  for item in generated_offer),
        'total_delivered':sum(item.quantity for item in products_shipped),
        'total_invoiced': sum(invoice.total_price  for invoice in invoices),
    }


    offer_info['total_remaining'] = offer_info['total_product'] - offer_info['total_delivered']

    if valor_total:
        offer_info['total_offer'] = valor_total.total_value
        offer_info['total_value_remaining'] = valor_total.total_value - offer_info['total_invoiced']
    else:
        offer_info['total_offer'] = 0
        offer_info['total_value_remaining'] = 0

    
    return offer_info


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def edit_offer(request,id):
    project = Project.objects.filter(id=id).first()
    tabs = project.tabs.all()
    generated_tab = GeneratedOffer.objects.filter(product__isnull = True, is_title= True, parent__isnull = False, project = project)
    generated_offer = GeneratedOffer.objects.filter(Q(project= project) and Q(product__isnull = False))
    subtotal = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",project= project).first()
    descuento = GeneratedOffer.objects.filter(measure__icontains="DESCUENTO",project= project).first()
    subtotal_descuento = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL MENOS (-) DESCUENTO",project= project).first()
    admin_descuento = GeneratedOffer.objects.filter(measure__icontains="ADMINISTRACIÓN",project = project).first()
    imprevistos = GeneratedOffer.objects.filter(measure__icontains="IMPREVISTROS",project = project).first()
    utilidades = GeneratedOffer.objects.filter(measure__icontains="UTILIDADES",project = project).first()
    subtotal_directo_indirecto = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL COSTO DIRECTO + INDIRECTO",project = project).first()
    iva_utilidad = GeneratedOffer.objects.filter(measure__icontains="IVA/UTILIDAD",project = project).first()
    valor_total = GeneratedOffer.objects.filter(measure__icontains="VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",project = project).first()
    remissions = Remission.objects.filter(project = project).all()    
    products_shipped = ProductSent.objects.filter(remission__in = remissions)
    generated_offer_ = GeneratedOffer.objects.filter(project= project, product__isnull = False)
    subtotals = GeneratedOffer.objects.filter(project = project, is_subtotal = True) 
    invoices = Invoice.objects.filter(project= project)
    folder = Folder.objects.filter(project = project).first()
    generated_offer__ = GeneratedOffer.objects.filter(project= project).exists()  
    if not generated_offer__:
        messages.info(request, "Asegurese de generar la oferta primero antes de editar!")

    
    # offer_info = {
    #     'total_product':sum(item.quantity  for item in generated_offer),
    #     'total_delivered':sum(item.quantity for item in products_shipped),
    #     'total_invoiced': sum(invoice.total_price  for invoice in invoices),
    # }


    # offer_info['total_remaining'] = offer_info['total_product'] - offer_info['total_delivered']
    
    # if valor_total:
    #     offer_info['total_value_remaining'] = valor_total.total_value - offer_info['total_invoiced']
    
    offer_info = calc_offer_info(project)

   
    products_info = [ ]


    for item in generated_offer_:
        dict_p = {}
        dict_p['product'] = item.product
        dict_p['parent'] = item.parent
        dict_p['delivered'] = 0
        dict_p['remaining'] = item.quantity 
        for product_sh in products_shipped:
            if item.section == product_sh.section:
                if item.tab == product_sh.tab:
                    if item.product == product_sh.product:
                        dict_p['product'] = item.product    
                        dict_p['parent'] = item.parent
                        dict_p['delivered'] = product_sh.quantity
                        dict_p['remaining'] = item.quantity - product_sh.quantity

            
        products_info.append(dict_p)
 
    sections = GeneratedOffer.objects.filter(product__isnull = True, is_title= True, parent__isnull = True, project = project)
    
    # 'remaining': item.quantity - next((product_sh.quantity for product_sh in products_shipped if product_sh.product.id == item.product.id), 0)
    

    return render(request,"edit_offer.html",{'project':project,
                                             'folder':folder,
                                             'generated_offer':generated_offer,
                                             'tabs':tabs,
                                             'subtotal':subtotal,
                                             'descuento':descuento,
                                             'subtotal_descuento':subtotal_descuento,
                                             'admin_descuento':admin_descuento,
                                             'imprevistos':imprevistos,
                                             'utilidades':utilidades,
                                             'subtotal_directo_indirecto':subtotal_directo_indirecto,
                                             'iva_utilidad':iva_utilidad,
                                             'valor_total':valor_total,
                                             'offer_info':offer_info,
                                             'products_shipped':products_shipped,
                                             'products_info':products_info,
                                             'generated_tab':generated_tab,
                                             'sections':sections,
                                             'subtotals':subtotals
                                             })

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_offer_item(request,id):
    referer = request.META.get("HTTP_REFERER")
    item = GeneratedOffer.objects.filter(id=id).first()
    if item:
        item.delete()
    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def save_offer_item(request):
    referer = request.META.get("HTTP_REFERER")
   
    if request.method == "POST":
        raw_data = request.body
        # print(f"desde save_item: {raw_data}"  )
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        
        subtotal = 0
        valor_descuento = 0
        subtotal_directo_indirecto_valor = 0
        subtotal_modules = 0
    
    
        for item in data['items']:
            offer_item = GeneratedOffer.objects.filter(id=item.get("id")).first()
            
            project = offer_item.project
            
            
            if item.get("title"):
                offer_item.title= item.get("title")

            if item.get("description"):
                offer_item.description= item.get("description")

            if item.get("measure"):
                offer_item.measure= item.get("measure")

            if item.get("quantity"):
                offer_item.quantity= int(item.get("quantity"))

            if item.get("unit_value"):
                offer_item.unit_value=  float(item.get("unit_value"))

            if item.get("total_value"):
                offer_item.total_value= float(item.get("total_value"))

            if offer_item.is_subtotal:
                subtotal += offer_item.total_value
            offer_item.save()
        
        try:

            subtotal_item = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",project= project).first()
            if subtotal_item:
                subtotal_item.total_value = subtotal
                subtotal_item.save()
            
            descuento_item = GeneratedOffer.objects.filter(measure__icontains="DESCUENTO",project= project).first()
            if descuento_item:
                descuento_item.porcent = data['descuento']
                valor_descuento = subtotal * (float(data['descuento'])/100)
                descuento_item.total_value = valor_descuento
                descuento_item.save() 
           

            subtotal_descuento = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL MENOS (-) DESCUENTO",project= project).first()
            
            if subtotal_descuento:
                subtotal_desc = subtotal - valor_descuento
                subtotal_descuento.total_value = subtotal_desc
                subtotal_descuento.save()

            
            admin_descuento = GeneratedOffer.objects.filter(measure__icontains="ADMINISTRACIÓN",project = project).first()
            
            if admin_descuento:
                admin_descuento.porcent = data['descuento_admin']
                descuento_admin_valor = subtotal_desc * (float(data['descuento_admin'])/100)
                admin_descuento.total_value = descuento_admin_valor
                admin_descuento.save()
            
           
            imprevistos = GeneratedOffer.objects.filter(measure__icontains="IMPREVISTROS",project = project).first()
            
            if imprevistos:
                imprevistos.porcent = data['descuento_imprevistos']
                descuento_imprevistos_valor = subtotal_desc * (float(data['descuento_imprevistos'])/100)
                imprevistos.total_value = descuento_imprevistos_valor
                imprevistos.save()

            
            utilidades = GeneratedOffer.objects.filter(measure__icontains="UTILIDADES",project = project).first()

            if utilidades:
                utilidades.porcent = data['descuento_utilidades']
                descuento_utilidades_valor = subtotal_desc * (float(data['descuento_utilidades'])/100)
                utilidades.total_value = descuento_utilidades_valor
                utilidades.save()
            

            subtotal_directo_indirecto = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL COSTO DIRECTO + INDIRECTO",project = project).first()

            if subtotal_directo_indirecto:
                
                subtotal_directo_indirecto_valor = subtotal_desc + descuento_admin_valor + descuento_imprevistos_valor + descuento_utilidades_valor
                subtotal_directo_indirecto.total_value = subtotal_directo_indirecto_valor
                subtotal_directo_indirecto.save()


            iva_utilidad = GeneratedOffer.objects.filter(measure__icontains="IVA/UTILIDAD",project = project).first()
            
            if iva_utilidad:
                iva_utilidad.porcent = data['iva_utilidad']
                iva_utilidad_valor = descuento_utilidades_valor * (float(data['iva_utilidad'])/100)
                iva_utilidad.total_value = iva_utilidad_valor
                iva_utilidad.save()
            

            valor_total = GeneratedOffer.objects.filter(measure__icontains="VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",project = project).first()
            
            if valor_total:
                valor_total_ = subtotal_directo_indirecto_valor + iva_utilidad_valor
                valor_total.total_value = valor_total_
                valor_total.save()

            

            return JsonResponse({'subtotal': round(subtotal,2), 
                                'valor_descuento': round(valor_descuento,2),
                                'subtotal_desc':round(subtotal_desc,2),
                                'descuento_admin_valor':round(descuento_admin_valor,2),
                                'descuento_imprevistos_valor':round(descuento_imprevistos_valor,2),
                                'descuento_utilidades_valor':round(descuento_utilidades_valor,2),
                                'subtotal_directo_indirecto_valor':round(subtotal_directo_indirecto_valor,2),
                                'iva_utilidad_valor':round(iva_utilidad_valor,2),
                                'valor_total':round(valor_total_,2),
                                'subtotal_modules':subtotal_modules
                                })
        except:
            
            messages.error(request,"No se ha podido guardar la oferta porque ha ocurrido un error asegurese de generar primero la oferta!")
            return redirect(referer)

            


        
    
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_from_car_to_offer(request,project_id,parent_id):

    referer = request.META.get("HTTP_REFERER")
    
    products_in_car = ProductBox.objects.filter(usersession = request.user).all() # productos en carrito

    project = Project.objects.filter(id = project_id).first() #consulta el proyecto

    offer_items = GeneratedOffer.objects.filter(project = project, product__isnull = False) # consulta los items que son productos
    total_value_ = 0

    for product_in_car in products_in_car: # recorre los elemetos del carrito
        total_value_ = product_in_car.quantity * product_in_car.price
        
        product_in_offer = GeneratedOffer.objects.filter(project = project, product__id = product_in_car.product.id, parent_id = parent_id).first()

        if product_in_offer:

            product_in_offer.quantity += product_in_car.quantity
            product_in_offer.save()

        else:    
            GeneratedOffer.objects.create(
                            product = product_in_car.product,
                            measure = product_in_car.product.measure,
                            description = product_in_car.product.description,
                            quantity = product_in_car.quantity,
                            unit_value = product_in_car.price,
                            total_value = total_value_,
                            project_id = project_id,
                            parent_id = parent_id
                            )
            
    messages.success(request,"El producto ha sido agregado correctamente")


    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def change_to_purcharse_order(request):
    if request.method == "POST":
        raw_data = request.body
        # print(f"desde save_item: {raw_data}"  )
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)

        item = GeneratedOffer.objects.filter(id = int(data['item_id']) ).first()
        item.to_purcharse_order = not item.to_purcharse_order
        item.save()
        return JsonResponse({'item_id':item.id, 'status':item.to_purcharse_order})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_order_from_offer(request, project_id): 
    project = Project.objects.filter(id = project_id).first()
    items_offer = GeneratedOffer.objects.filter(project = project, to_purcharse_order = True)
    
    ProductBox.objects.filter(usersession = request.user).all().delete()
    
    for item in items_offer:
        if item.to_purcharse_order:
            product = item.product
            quantity = item.quantity
            price = item.unit_value
            usersession = request.user
            ProductBox.objects.create(product = product, 
                                      quantity = quantity, 
                                      price=price, 
                                      usersession = usersession)
            
    return redirect(f"/createorder/{project_id}")



@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_remission_from_offer(request, project_id): 

    project = Project.objects.filter(id = project_id).first()
    items_offer = GeneratedOffer.objects.filter(project = project, to_purcharse_order = True)
    
    ProductBox.objects.filter(usersession = request.user).all().delete()

    
    for item in items_offer:
        if item.to_purcharse_order:
            product = item.product
            quantity = item.quantity
            price = item.unit_value
            usersession = request.user
            ProductBox.objects.create(product = product, 
                                      quantity = quantity, 
                                      price=price,
                                      tab = item.tab,
                                      section = item.section, 
                                      usersession = usersession)
            
    return redirect(f"/createremission/{project_id}")



@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_car_to_invoice(request,id):
    referer = request.META.get("HTTP_REFERER")
    productbox = ProductBox.objects.filter(usersession = request.user).all()
    invoice = Invoice.objects.filter(id = id).first()

    # for productb in productbox:
    #         if productb.quantity > productb.product.quantity:
    #             messages.error(request,f"No hay sufiente cantidad del producto {productb.product.model} en el inventario")
    #             return redirect(referer)
    # products = []
    for productb in productbox:
        products_invoice = ProductInvoice.objects.create(
            product = productb.product,
            quantity = productb.quantity,
            price = productb.price,
            invoice = invoice
        )
        # product = Product.objects.filter(id= productsent.product.id).first()
        # products.append(product)
        # product.quantity -= productsent.quantity
        # product.save()
    
    # seve_stat_prod(products)

    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_product_in_invoice(request,id):  
    
    referer = request.META.get("HTTP_REFERER")
    
    remissionproduct = ProductInvoice.objects.filter(id = id).first()


    remissionproduct.delete()

    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_invoice_from_offer(request, project_id): 

    project = Project.objects.filter(id = project_id).first()
    items_offer = GeneratedOffer.objects.filter(project = project, to_purcharse_order = True)
    
    ProductBox.objects.filter(usersession = request.user).all().delete()

    
    for item in items_offer:
        if item.to_purcharse_order:
            product = item.product
            quantity = item.quantity
            price = item.unit_value
            usersession = request.user
            ProductBox.objects.create(product = product, 
                                      quantity = quantity, 
                                      price=price, 
                                      usersession = usersession)
            
    return redirect(f"/createinvoice/{project_id}")


@permission_required('tsinc.view_orderinvoice', login_url='/accessdenied/')
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def show_all_purcharse_order_invoices(request):

    order_invoices = OrderInvoice.objects.all()

    return render(request,"purcharse_order_invoices.html",{'order_invoices':order_invoices}) 


    
@login_required
def download_categories(request):
    # Consulta todos los datos del modelo
    queryset = Category.objects.all().values()
    # Convierte el queryset a un DataFrame de pandas
    df = pd.DataFrame(queryset)

    # Crear una respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tabla_categorias.xlsx'

    # Guardar el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Categorias')
    return response


@user_passes_test(staff_required,login_url='/accessdenied/')    
@login_required   
def upload_categories(request):
    referer = request.META.get("HTTP_REFERER")
    if request.method == 'POST':
        form = UploadFile(request.POST, request.FILES)
    
        if form.is_valid():
            excel_file = request.FILES['file']
            
            df = pd.read_excel(excel_file, engine='openpyxl')

            for _,row in df.iterrows():
                id = int(row['id'])
                name = row['name']
                tag = row['tag']
                parent_id = row['parent_id'] if pd.notna(row['parent_id']) else None
            

                category = Category.objects.filter(id= id).first()
                
                
                if category:
                    category.name = name
                    category.tag = tag
                    category.parent_id = parent_id
                    category.save()
                else:
                        Category.objects.create(
                            name = name,
                            tag = tag,
                            parent_id = parent_id, 
                        )


            messages.success(request, 'Categorias cargadas correctamente.')
            return redirect(referer)
        else: 
            messages.error(request, 'Las categorias no se cargaron correctamente.')    
            return redirect(referer)
    else:
        form = UploadFile()
    return render(request, 'upload_categories.html', {'form': form})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def create_task(request, project_id):

    referer = request.META.get("HTTP_REFERER")

    project = Project.objects.filter(id= project_id).first()
    folder = Folder.objects.filter(project = project).first()

    if request.method == "GET":
        return render(request,"create_task.html", {'form':CreateTask(), 'folder':folder}) 
    else:
        name = request.POST["name"]
        start_date = request.POST["start_date"]
        due_date = request.POST["due_date"]
        description = request.POST["description"]
        users = request.POST.getlist("users")
       
        task = Task.objects.create(
            name=name,
            start_date=start_date,
            due_date=due_date,
            description=description,
            project = project,
            state = 'pendiente',
            container = 'container1',
            assigned_by = request.user

        )
        # Asignar los usuarios a la tarea
        task.users.set(users)
        task.save()
        
        for user in users:
            # send_notification(task,user.email)
            user_ = User.objects.filter(id = user).first()
            if request.POST.getlist("send_email"):
                if (user_.email):
                    send_notification(task,user_.email)
                    messages.success(request,f"Se ha enviado una notificacion al emaíl del usuario {user_}")
                else:
                    messages.info(request,f"El usuario {user_} no tiene un correo asignado, porfavor verificar en el panel de adminitrador")

        
        calculate_and_save_percentage_of_tasks(project)
        messages.success(request,"La tarea ha sido creada correctamente")
    
        return redirect(referer)


def edit_task(request, id):
    task = get_object_or_404(Task, id=id)
    if request.method == 'POST':
        referer = request.META.get("HTTP_REFERER")
        form = CreateTask(request.POST)
        if form.is_valid():
            task.name = form.cleaned_data['name']
            task.start_date = form.cleaned_data['start_date']
            task.due_date = form.cleaned_data['due_date']
            task.description = form.cleaned_data['description']
            task.save()  # Guarda la instancia
            # Actualiza la relación muchos a muchos
            task.users.set(form.cleaned_data['users'])
            messages.success(request,"La tarea ha sido actulizada correctamente")

            return redirect(referer)
              # Redirige a la lista de tareas o a donde desees
    else:
        initial_data = {
            'name': task.name,
            'start_date': task.start_date,
            'due_date': task.due_date,
            'description': task.description,
            'users': task.users.all() 
        }
        form = CreateTask(initial=initial_data)
    return render(request, 'create_task.html', {'form': form})

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def custom_offer(request):
    products = list(Product.objects.values())  # Convertir queryset a lista de diccionarios

    return render(request,"custom_offer.html",{'products':products}) 



@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def save_offer_titles(request):
    if request.method == "POST":
        raw_data = request.body
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        return JsonResponse({'status':'ok'})

def obtener_productos(request):
    productos = list(Product.objects.values())
    return JsonResponse(productos, safe=False)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_section(request,project_id):
    referer = request.META.get("HTTP_REFERER")

    item = GeneratedOffer.objects.filter(project_id=project_id).order_by('-section').first()


    new_section = GeneratedOffer.objects.create(title=f"NUEVA SECTION",project_id = project_id, is_title = True)
    GeneratedOffer.objects.create(title="NUEVO SUBTITULO", project_id = project_id, is_title = True, parent = new_section )

    GeneratedOffer.objects.create(measure="SUBTOTAL", project_id = project_id, is_subtotal = True, parent = new_section )

    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_offer_item(request,id):
    referer = request.META.get("HTTP_REFERER")
    
    item = GeneratedOffer.objects.filter(id=id).first()
    item.delete()

    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def add_subtitle(request,project_id, parent_id):
    referer = request.META.get("HTTP_REFERER")
    
    GeneratedOffer.objects.create(title="NUEVO SUBTITULO",project_id = project_id, parent_id = parent_id, is_title = True)

    return redirect(referer)


@user_passes_test(superuser_required,login_url='/accessdenied/')
@login_required
def show_all_activity(request):
    
    all_activiy = Activity.objects.filter().order_by('-date').all()
    search = request.GET.get('search')
    if search:
        all_activiy = Activity.objects.filter(Q(model__icontains= search) | Q(usersession__username__icontains= search))
    
    all_activiy = paginator(request,all_activiy,10)

    return render(request,"show_all_activity.html",{ 
        'all_activity':all_activiy
    })

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def user_view(request):

    tasks = Task.objects.filter(users = request.user)
    projects = set([ task.project for task in tasks ])

    all_archived = False
    
    tasks_per_project = [
        {
            'project':project,
            'tasks': [task for task in tasks if project.id == task.project.id],
            'total': sum(1 for task in tasks if project.id == task.project.id)
        }

    for project in projects

    ]


    all_archived = all([True if project.archive_tasks else False for project in projects ])

    
    
    return render(request,"user_view.html", {'tasks_per_project':tasks_per_project, 'all_archived':all_archived})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def tasks(request,project_id):

    project = Project.objects.filter(id=project_id).first()

    tasks = Task.objects.filter(project = project, users = request.user)

    return render(request,"tasks.html", {'tasks':tasks})



def upload_task_file(request):
    if request.method == 'POST':
        # file = request.FILES['file']
    
        if 'file' in request.FILES:
            file = request.FILES['file']
            task_id = request.POST.get('task_id')

            file_without_space = file.name.replace(" ", "_")
            
            task_file = File.objects.filter(name=file_without_space,task_id = task_id).first()

            task = Task.objects.filter(id = task_id).first()

            folder_name = "task_files"

            if task_file:

                upload_file(file,file_without_space,task_file.path)

            else:
                task_file = File.objects.create(name = file_without_space,
                                                            path=f"{folder_name}/{task.id}",
                                                            task_id=task.id,
                                                            usersession = request.user
                                                            )  
            
            upload_file(file,file_without_space,task_file.path)

        return JsonResponse({'message': 'File uploaded successfully'})
    
    return JsonResponse({'error': 'No file uploaded'}, status=400)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def save_order_container_task(request):
    if request.method == "POST":
        raw_data = request.body
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)

        task = Task.objects.filter(id = data['task_id']).first()


        task.container = data['container']

        if task.container == "container1":
            task.state = 'pendiente'
        elif task.container == "container2":
            task.state = 'en proceso'
        elif task.container == "container3":
            task.state = 'finalizado'

        task.save()
        calculate_and_save_percentage_of_tasks(task.project)

        return JsonResponse({'status':'ok'})


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def save_comment(request):
    if request.method == "POST":
        raw_data = request.body
        body_unicode = raw_data.decode('utf-8')
        data = json.loads(body_unicode)
        task = Task.objects.filter(id = data['task_id']).first()
        if data['message']:
            Comment.objects.create(message = data['message'], task = task, usersession = request.user )
            return JsonResponse({'status':'ok'})
        else:
            return JsonResponse({'status':'error'})
           
            
@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_task(request,id):  
    referer = request.META.get('HTTP_REFERER')
    task = Task.objects.filter(id = id).first()
    task.delete()
    return redirect(referer)


@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def delete_comment(request,id):  
    referer = request.META.get('HTTP_REFERER')
    comment = Comment.objects.filter(id = id).first()
    comment.delete()
    return redirect(referer)

@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def archive_task(request,id,opt):  
    referer = request.META.get('HTTP_REFERER')
    if opt == 0:
        project = Project.objects.filter(id = id).first()
        project.archive_tasks = True
        project.save()

    elif opt == 1:
        project = Project.objects.filter(id = id).first()
        project.archive_tasks = False
        project.save()

    return redirect(referer)



@user_passes_test(staff_required,login_url='/accessdenied/') 
@login_required
def archived_task_project(request):

    tasks = Task.objects.filter(users = request.user)
    projects = set([ task.project for task in tasks ])
    
    tasks_per_project = [
        {
            'project':project,
            'tasks': [task for task in tasks if project.id == task.project.id],
            'total': sum(1 for task in tasks if project.id == task.project.id)
        }

    for project in projects

    ]

    all_archived = any([True if project.archive_tasks else False for project in projects ])

    return render(request,"archived_task_project.html", {'tasks_per_project':tasks_per_project, 'all_archived':all_archived})


class ProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return Product.objects.none()  # Evita mostrar resultados si el usuario no está autenticado.

        qs = Product.objects.all()

        if self.q:
            qs = qs.filter(model__icontains=self.q)  # Filtra productos por el nombre (o el campo que prefieras).

        return qs