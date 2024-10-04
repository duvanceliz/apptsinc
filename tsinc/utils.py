from django.urls import reverse
from .models import *
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import functools
import time
import math
import openpyxl
from collections import Counter
import random
from datetime import date
from openpyxl.drawing.image import Image
import os
from django.conf import settings




thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )

def print_row(sheet,data,start_row,column,config=None): 
        if config is None:
            config = {
                'font': False,
                'bold': False,
                'font_color': "000000",
                'border': False,
                'alig': False,
                'wrap': False,
                'h':"left",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
            
        for i in range(0,len(data)):
            element = data[i]
            cell = sheet.cell(row=i+start_row, column=column, value=element)

            if config.get('font', False):
                cell.font = Font(bold=config.get('bold',False), color=config.get('font_color', "000000"))

            if config.get('border', False):
                cell.border = thin_border

            if config.get('alig', False):
                cell.alignment = Alignment(wrap_text=config.get('wrap',False),horizontal=config.get('h',"left"), vertical=config.get('v',"center"))


            if config.get('fill', False):
                cell.fill = PatternFill(start_color=config.get('fill_color', "FFFFFF"), end_color=config.get('fill_color', "FFFFFF"), fill_type="solid")
    
def print_column(sheet,data,row,start_column,config=None):
    if config is None:
            config = {
                'font': False,
                'bold': False,
                'font_color': "000000",
                'border': False,
                'alig': False,
                'wrap': False,
                'h':"left",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    
        
    for i in range(0,len(data)):
        element = data[i]
        cell = sheet.cell(row=row, column=i+start_column, value=element)
        if config.get('font', False):
            cell.font = Font(bold=config.get('bold',False), color=config.get('font_color', "000000"))

        if config.get('border', False):
            cell.border = thin_border

        if config.get('alig', False):
            cell.alignment = Alignment(wrap_text=config.get('wrap',False),horizontal=config.get('h',"left"), vertical=config.get('v',"center"))

        if config.get('fill', False):
            cell.fill = PatternFill(start_color=config.get('fill_color', "FFFFFF"), end_color=config.get('fill_color', "FFFFFF"), fill_type="solid")

def closest_match(input_one,input_two):

    def generar_combinaciones(num_ui, num_co, bo, bi, ao):
        combinaciones = []
        # Combinaciones de UI (longitud num_ui)
        for ai_count in range(num_ui + 1):  # 0 to num_ui
            bi_count = num_ui - ai_count
            # Combinaciones de CO (longitud num_co)
            for ao_count in range(num_co + 1):  # 0 to num_co
                bo_count = num_co - ao_count
                combinacion = [ai_count, bi_count + bi, ao_count + ao, bo_count + bo]
                combinaciones.append(combinacion)
        return combinaciones


    # Paso 2: Calcular la distancia entre combinaciones
    def calcular_distancia(combinacion, entrada):
        distancia = 0
        for i in range(len(combinacion)):
            distancia += abs(combinacion[i] - entrada[i])
        return distancia

    # Paso 3: Encontrar la combinación más cercana
    def encontrar_combinacion_mas_cercana(combinaciones, entrada):
        mejor_combinacion = None
        menor_distancia = float('inf')
        for combinacion in combinaciones:
            distancia = calcular_distancia(combinacion, entrada)
            if distancia < menor_distancia:
                menor_distancia = distancia
                mejor_combinacion = combinacion
        return mejor_combinacion


    # Generar combinaciones desde el arreglo con UI y CO
    num_ui = input_one[0]  # Total UI
    num_co = input_one[4]  # Total CO
    bo = input_one[3]  # Total BO
    bi = input_one[1]
    ao = input_one[2]

    combinaciones_posibles = generar_combinaciones(num_ui, num_co,bo,bi,ao)

    return encontrar_combinacion_mas_cercana(combinaciones_posibles, input_two)

def sort_list_point(points):
    point_list = [0]*5
    for point in points:
        if len(point) == 3:
            point = f"0{point}"   
        if point[2:4] == "AI" or point[2:4] == "UI":
            point_list[0] = int(point[:2])
        elif point[2:4] == "BI" or point[2:4] == "DI":
            point_list[1] = int(point[:2])
        elif point[2:4] == "AO":
            point_list[2] = int(point[:2])
        elif point[2:4] == "BO" or point[2:4] == "DO":
            point_list[3] = int(point[:2]) 
        elif point[2:4] == "CO":
            point_list[4] = int(point[:2]) 
    return point_list 

def clean_points(points):
    return [max(point,0) for point in points]

def subtract_points(a, b):
    return [x - y for x, y in zip(a, b)]

def add_points(*args):
    return [sum(values) for values in zip(*args)]

def find_expansion(total_points,controller_and_expansions,expansions):
    def calc_distance(input_one,input_two):
        distance = 0
        for i in range(len(input_one)):
            distance += abs(input_one[i] - input_two[i])
        return distance
    
    expansion_choice = None
    shoter_distance = float('inf')
    points = [ item['points'] for item in controller_and_expansions]
    add_p = add_points(*points)
    for expansion in expansions:
        cller_exp_point = add_points(expansion['points'],add_p) 
        # print(f"{expansion['points']}+{add_p} result:{cller_exp_point}")
        new_cller_exp_point = closest_match(cller_exp_point,total_points)
        expansion['ce_basic_points'] = new_cller_exp_point
        distance = calc_distance(new_cller_exp_point, total_points)
        #print(total_points, f"todo:{new_cller_exp_point}{expansion['model']}:{distance}")
        if distance < shoter_distance:
            shoter_distance = distance
            expansion_choice = expansion
    return expansion_choice
        
def calc_controllers(item_points, type_controller):
    
    def iterator_function(total_points,filtered_expansions):
        expansion = find_expansion(total_points,controller_and_expansions,filtered_expansions)
        controller_and_expansions.append(expansion)
        new_points = subtract_points(total_points, expansion['ce_basic_points'])
        return new_points
    
    def filter_expansions(total_points,expansions):
        prom = 10
        total_points = clean_points(total_points)
        reduce_points = functools.reduce(lambda x,y:x+y,total_points)
        if reduce_points <= prom:
            filtered_expansions = [expansion for expansion in expansions if expansion['total_points'] <= prom]
        elif reduce_points > prom:
            filtered_expansions = [expansion for expansion in expansions if expansion['total_points'] > prom]
        return filtered_expansions
        
    
    def add_expansion_js(total_points,expansions):
        stop=0
        filtered_expansions = filter_expansions(total_points, expansions)
        new_points = iterator_function(total_points,filtered_expansions)
        while any( point > 0 for point in new_points):
            stop +=1
            filtered_expansions = filter_expansions(new_points, expansions)
            new_points = iterator_function(total_points,filtered_expansions)
            if stop > 10:
                print(new_points)
                break
                
     
            
    def add_expansion_controllers(new_points, expansions):
        for expansion in expansions:
            if expansion['model'] == 'VMQ-02D13':
                add_expansion(expansion, new_points, 3, 8)
            elif expansion['model'] == 'PJ-E08A1':
                add_expansion(expansion, new_points, 2, 8)
            elif expansion['model'] == 'VMQ-30U13':
                add_expansion_ui14(expansion, new_points, 0, 1, 14)
    

    def add_expansion(expansion, new_points, index, limit):
        while new_points[index] > 0:
            controller_and_expansions.append(expansion)
            new_points[index] -= limit

    def add_expansion_ui14(expansion, new_points, index1, index2, limit):
        new_points = clean_points(new_points)
        sum_bi_ai = new_points[index1] + new_points[index2]
        while sum_bi_ai > 0:
            controller_and_expansions.append(expansion)
            sum_bi_ai -= limit
    
    def find_cllers_exp(cller_type):
         
        if cller_type == "JCI FACILITY EXPLORER":
            controllers = Product.objects.filter(code="C-FE")
            expansions = Product.objects.filter(code='E-FE')
        else:
            controllers = Product.objects.filter(code="C-M")
            expansions = Product.objects.filter(code='E-M')
        # print(expansions)
       
        controllers_js = [{
            'is_controller':True,
            'model':controller.model,
            'points':sort_list_point(controller.point.split("-")) 
        }
        for controller in controllers   
        ]
        expansions_js = [{
            'is_controller':False,
            'model':expansion.model,
            'points':sort_list_point(expansion.point.split("-")),
            'total_points': functools.reduce(lambda x,y: x+y,sort_list_point(expansion.point.split("-")))
        }
        for expansion in expansions  
        ]   
        return  controllers_js, expansions_js  

    controller_and_expansions = []
    if type_controller == 'LG BECON CONTROLLER':
        controller = Product.objects.get(code='C-LG')
        expansions = Product.objects.filter(code__icontains='E-LG')
        controller_lg = {
            'is_controller':True,
            'model':controller.model,
            'points':sort_list_point(controller.point.split("-"))
        }
        expansions_lg = [{
            'is_controller':False,
            'model':expansion.model,
            'points':sort_list_point(expansion.point.split("-")) 
        }
        for expansion in expansions   
        ]

        point_quantity = functools.reduce(lambda x,y: x+y,item_points)

        if point_quantity > 0:
            controller_points = controller_lg['points']
            if all(item <= controller for item, controller in zip(item_points, controller_points)):
                controller_and_expansions.append(controller_lg)
            else:
                controller_and_expansions.append(controller_lg)
                new_points = subtract_points(item_points, controller_points)
                add_expansion_controllers(new_points, expansions_lg)
    
    elif type_controller == 'JCI FACILITY EXPLORER' or type_controller == 'JCI METASYS':
      
        controllers_js, expansions_js = find_cllers_exp(type_controller)
        
        if len(controllers_js) > 0 and len(expansions_js) > 0:

            point_quantity = functools.reduce(lambda x,y: x+y,item_points)
        
            if point_quantity > 0 and point_quantity < 18:
                controller = [controller for controller in controllers_js if controller['model']=="F4-CGM04060" or controller['model']=="M4-CGM04060"][0]  
                new_controller_point = closest_match(controller['points'],item_points)
                
                #print(f'puntos controlador:{new_controller_point}')
                if all(item <= controller for item, controller in zip(item_points, new_controller_point)):
                    controller_and_expansions.append(controller)
                else:
                    controller_and_expansions.append(controller)
                    # new_points = subtract_points(item_points, new_controller_point)
                    add_expansion_js(item_points,expansions_js)
            else:
                controller = [controller for controller in controllers_js if controller['model']=="F4-CGM09090" or controller['model']=="M4-CGM04060"][0]  
                new_controller_point = closest_match(controller['points'],item_points)
                
                # print(f'puntos controlador:{new_controller_point}')
                if all(item <= controller for item, controller in zip(item_points, new_controller_point)):
                    controller_and_expansions.append(controller)
                else:
                    controller_and_expansions.append(controller)
                    # new_points = subtract_points(item_points, new_controller_point)
                    add_expansion_js(item_points,expansions_js)


    return controller_and_expansions

def print_title(sheet,title,coor,opt):
    if opt == 1:
        cell = sheet[coor]
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAD7D7", end_color="DAD7D7", fill_type="solid")
        cell.border = thin_border
    elif opt == 2:
        cell = sheet[coor]
        cell.value = title
        cell.font = Font(bold=True)
        cell.border = thin_border
    elif opt == 3:
        cell = sheet[coor]
        cell.value = title
        cell.font = Font(bold=True)

def count_point(sheet,col):
    sum = 0
    for cell in sheet[col]:
        if isinstance(cell.value, (int, float)):  # Asegurarse de que los valores sean numéricos
            sum += cell.value
    return sum

def count_point_column(sheet,row, start, end):
    total_sum = 0
    for col in range(start, end):  # M es la columna 13 y Q es la columna 17
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):  # Asegurarse de que los valores sean numéricos
            total_sum += cell.value
    return total_sum


def count_value_row(sheet,col, start, end):
    total_sum = 0
    for row in range(start, end):  # M es la columna 13 y Q es la columna 17
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):  # Asegurarse de que los valores sean numéricos
            total_sum += cell.value
    return total_sum

def print_data(data,tab,sheet,project):
    
    sheet_model = Sheet.objects.create(name = sheet.title, project = project, tab=tab)
    divices = Divice.objects.filter(tab = tab, project = project).all()
    divices.delete()
    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"left",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    
    config_style_titles2 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"left",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    config_style_titles3 = {
                            'font':False,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"left",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }

    print_row(sheet,['PUNTOS DISPONIBLES EN CONTROLES',
                     'PUNTOS PROYECTADOS',
                     'PUNTOS SOBRANTES'],4,8,config_style_titles1)

    print_row(sheet,['<==',
                     '<==',
                     '<=='],4,7,config_style_titles2)

    print_row(sheet,['PROYECTO',
                     'SISTEMA',
                     'SOLUCIÓN'],2,1,config_style_titles1)


    print_column(sheet,['ITEM',
                        'DESCRIPCIÓN',
                        'EA',
                        'ED',
                        'SA',
                        'SD',
                        None,
                        'NOTAS / TIPO DE SENSOR / ACTUADOR', 
                        ],8,1,config_style_titles1)
    
    print_row(sheet,[project.name.upper(),
                     sheet.title,
                     tab.controller],2,2,config_style_titles2)

    
    cell_dimension = ['C','D','E','F']

    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['H'].width = 50
    sheet.column_dimensions['G'].width = 4
    sheet.column_dimensions['L'].width = 32
    sheet.column_dimensions['I'].width = 2
    sheet.column_dimensions['J'].width = 2
    sheet.column_dimensions['K'].width = 5

    for cell in cell_dimension:
        sheet.column_dimensions[f'{cell}'].width = 5
        

    units_per_tab = list(filter(lambda dict: dict['tab_name'] == f'{sheet.title}',data))

    # print(units_tab)


    len_item = 0  
  
    for dict in units_per_tab:
        count = 0
        cell = sheet.cell(row=9+len_item, column=2, value=dict.get('unit_name'))
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        for i,item in enumerate(dict['related_items'], start=10+len_item):
            count+=1
            cell_index = sheet.cell(row=i, column=1, value=count)
            config_style_titles2['h'] = "center"
            style_cells([cell_index],config_style_titles2)

            if item.tag != 'None':
                cell_product_name = sheet.cell(row=i, column=2, value=item.tag)
            else:
                cell_product_name = sheet.cell(row=i, column=2, value=item.img.product.product_name)
            cell_description = sheet.cell(row=i, column=8, value= f'{item.img.product.model} - {item.img.product.brand}')
            
            if item.img.product.model != "AI" and item.img.product.model != "BI" and item.img.product.model != "AO" and item.img.product.model != "BO":
                Divice.objects.create(tag=item.tag, model = item.img.product.model, brand = item.img.product.model, tab = tab, project = project)
            
            points = item.img.product.point.split("-")
            for c,point in enumerate(sort_list_point(points)[0:4], start=3):
                if point != 0: 
                    cell = sheet.cell(row=i, column=c, value=point)
                    cell.font = Font(bold=True, color="FF0000")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                     
            cell_product_name.border = thin_border
            cell_description.border = thin_border
            cell_description.fill = PatternFill(start_color="FEFE7A", end_color="FEFE7A", fill_type="solid")
            

        len_item += len(dict.get('related_items')) + 1 
            



    points = [count_point(sheet,"C"),
              count_point(sheet,"D"),
              count_point(sheet,"E"),
              count_point(sheet,"F")]
    
    

    Points.objects.create(name="TOTAL_POINTS",eu=points[0],
                        ed=points[1],
                        sa=points[2],
                        sd=points[3],
                        sheet=sheet_model)



    print_column(sheet,points,5,3,config_style_titles2)
    

    cllers = calc_controllers(points, tab.controller)
   

    print_points_cller_exp_trf(sheet,points,cllers,sheet_model,project)


def get_code():
    return '{:04d}'.format(random.randint(0, 9999))


def calc_chest(c_quantity):
    if c_quantity > 0 and c_quantity <= 1:
        chest = Product.objects.filter(code="CO5050").first()
    elif c_quantity == 2:
        chest = Product.objects.filter(code="CO6060").first()
    elif c_quantity > 2:
        chest = Product.objects.filter(code="CO6080").first()
    else:
        chest = Product.objects.filter(code="CO5050").first()
    return chest    
    


def print_points_cller_exp_trf(sheet,points,cllers,sheet_model,project,modify=False):

    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    
    config_style_titles2 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    
    config_style_titles_fc_g = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"90D473"
                        }
   
    points_header_cllers_exp = ['TIPO','CONTROLADOR / EXPANSIÓN','EU','ED','SA','SD','SC','TOTAL']
    
    print_column(sheet,points_header_cllers_exp,8,11,config_style_titles1)

    c_quantity = 0

    if cllers:
        for i,controller in enumerate(cllers, start=9):
            c_quantity += controller['is_controller']
            cell = sheet.cell(row=i, column=12, value=controller['model'])
            cell.font = Font(bold=False)
            if not modify:
                Points.objects.create(name=controller['model'],
                                    is_controller = controller["is_controller"],
                                    eu=controller['points'][0],
                                    ed=controller['points'][1],
                                    sa=controller['points'][2],
                                    sd=controller['points'][3],
                                    sc=controller['points'][4],
                                    sheet=sheet_model) 
            cell.border = thin_border
            for y, point in enumerate(controller['points'],start=13):
                if point != 0:
                    cell = sheet.cell(row=i, column=y, value=point)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                    cell.border = thin_border
          
        for i in range(9, 9+len(cllers)):
            sum_point = count_point_column(sheet,i,13,18)
            cell = sheet.cell(row=i, column=18, value=sum_point)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = thin_border
    
    chest = calc_chest(c_quantity)
    if not modify:
        
        Divice.objects.create(tag=chest.code,
                              model=chest.model,
                              brand=chest.product_name,
                              tab = Tabs.objects.filter(tab_name=sheet.title,project=project).first(),
                              project = project
                              )
     

    row1 = 8 + len(cllers) + 2
    row2 = 8 + len(cllers) + 3
    row3 = 8 + len(cllers) + 4

    config_style_titles1['h'] = "left"
    print_row(sheet,['PUNTOS DISPONIBLES EN CONTROLES',
                     'PUNTOS PROYECTADOS',
                     'PUNTOS SOBRANTES'
                     ],row1,12,config_style_titles1)
    
    
    trafo_title = ["CANT","ACCESORIO"]
    print_column(sheet,trafo_title,row3+3,11,config_style_titles1)

    type_module = [ "C" if cller['is_controller'] else "E" for cller in cllers]

    print_row(sheet,type_module,9,11,config_style_titles2)

    points_cller_exp = [count_point(sheet,"M"),
                count_point(sheet,"N"),
                count_point(sheet,"O"),
                count_point(sheet,"P"),
                count_point(sheet,"Q")]

    print_column(sheet,points_cller_exp,row1,13,config_style_titles2)

    new_points_cller_exp = closest_match(points_cller_exp,points)

    print_column(sheet,new_points_cller_exp,4,3,config_style_titles2)

            
    leftover_points = [ x-y for x,y in zip(new_points_cller_exp,points) ]

    print_column(sheet,leftover_points,6,3,config_style_titles_fc_g)

    pp = [0]*5
    result = 0
    result2 = 0
    result3 = 0
 
    for i in range(1,4):
        re = 0
        if points[i] > points_cller_exp[i]:
            pp[i] = points_cller_exp[i]
            re = points[i] - points_cller_exp[i]
        else:
            pp[i] = points[i]
        if i == 1:
            result = re
        elif i == 2:
            result2 = re
        elif i == 3:
            result3 = re
    
    pp[0] = points[0] + result
    pp[4] = result2 + result3


    print_column(sheet,pp,row2,13,config_style_titles2)

    ps = subtract_points(points_cller_exp,pp)
    


    print_column(sheet,ps,row3,13,config_style_titles_fc_g)

   
    total_points_divice = 0 
    # print(sheet)
    for i in range(row1, row1+3):
        sum_point = count_point_column(sheet, i,13,18)
        if i == row1+1:
            total_points_divice += sum_point
        cell = sheet.cell(row=i, column=18, value=sum_point)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    total_point = Total_points.objects.filter(sheet=sheet.title, project = sheet_model.project).exists()
    
    if total_point:
        total_point = Total_points.objects.filter(sheet = sheet.title, project = sheet_model.project).first()
        total_point.sheet = sheet.title
        total_point.total = total_points_divice
    else:
       Total_points.objects.create(sheet = sheet.title, total=total_points_divice, project = sheet_model.project) 
    

    total_va = points[3] + len(cllers) * 14
    
    dble_total_va = total_va * 2
    
    trfs = {
        'tr40':'TR40VA004',
        'tr50':'TR50VA004',
        'tr75':'TR75VA004',
        'tr100':'TR100VA004',
    }

    if dble_total_va <= 40: 
        trf = trfs.get('tr40',None)
        quantity = 1
    elif dble_total_va > 40 and dble_total_va <= 50:
        trf = trfs.get('tr50',None)
        quantity = 1
    elif dble_total_va > 50 and dble_total_va <= 75:
        trf = trfs.get('tr75',None)
        quantity = 1
    elif dble_total_va > 75 and dble_total_va <= 100:
        trf = trfs.get('tr100',None)
        quantity = 1
    elif dble_total_va > 100:
        trf = trfs.get('tr100',None)
        quantity = math.floor(dble_total_va / 100)
        
    quantity_trf = [quantity,trf]
    
    if not modify:
        trf = Product.objects.filter(model = trf).first()
        tab = Tabs.objects.filter(tab_name = sheet.title, project=project).first()
        for i in range(1,quantity+1):
            Divice.objects.create(model=trf.model,brand = trf.brand, tab = tab, project = project)
    else:
        trf = Product.objects.filter(model = trf).first()
        tab = Tabs.objects.filter(tab_name = sheet.title, project= project).first()
        divice = Divice.objects.filter(model=trf.model, tab = tab, project = project).all()
        divice.delete()
        for i in range(1,quantity+1):
            Divice.objects.create(model=trf.model,brand = trf.brand, tab = tab, project = project)
    
    print_column(sheet,quantity_trf,row3+4,11,{
        'font':False,
        'font_color':"000000",
        'border':True,
        'alig':True,
        'fill':False,
        'fill_color':"90D473"
    })



def calc_trafo(sheet, row, column, divices,tab,project, points = 0):
    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    config_style_titles2 = {
                            'font':False,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': True,
                            'h':"center",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    total_va = points + len(divices) * 14
    
    dble_total_va = total_va * 2
    
    trfs = {
        'tr40':'TR40VA004',
        'tr50':'TR50VA004',
        'tr75':'TR75VA004',
        'tr100':'TR100VA004',
    }

    if dble_total_va <= 40: 
        trf = trfs.get('tr40',None)
        quantity = 1
    elif dble_total_va > 40 and dble_total_va <= 50:
        trf = trfs.get('tr50',None)
        quantity = 1
    elif dble_total_va > 50 and dble_total_va <= 75:
        trf = trfs.get('tr75',None)
        quantity = 1
    elif dble_total_va > 75 and dble_total_va <= 100:
        trf = trfs.get('tr100',None)
        quantity = 1
    elif dble_total_va > 100:
        trf = trfs.get('tr100',None)
        quantity = math.floor(dble_total_va / 100)

    tr = Divice.objects.filter(model=trf,tab=tab,project=project).all()
    tr.delete()
    tr_prod = Product.objects.filter(model = trf).first()
    for i in range(1,quantity+1):
        Divice.objects.create(model=tr_prod.model, brand = tr_prod.brand, tab = tab, project = project)
    
    
    sheet.column_dimensions['G'].width = 32
    
    print_column(sheet,["CANT","ACCESORIO"],row,column,config_style_titles1)
        
    print_column(sheet,[quantity,trf],row+1,column,config_style_titles2)   
   

def print_calc_supervirsor(sheet,tab,c_quantity,project):

    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    config_style_titles2 = {
                            'font':False,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': True,
                            'h':"center",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    
    sheet_model = Sheet.objects.create(name= sheet.title, project = tab.project, tab=tab )
    
    total_points_model = Total_points.objects.filter(project = project).all()
        
    total_points_divice = 0

    for total_p in total_points_model:
        total_points_divice += total_p.total

    supervisor_title = ['SUPERVISOR','DESCRIPCIÓN']
    title = ["LICENCIA","DESCRIPCIÓN"]
    
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 100
    
        
    print_column(sheet,supervisor_title,2,2,config_style_titles1)
    print_column(sheet,title,5,2,config_style_titles1)
    
    def save_licenses(licenses):
        sheet_model = Sheet.objects.filter(name="Supervisor", project = tab.project, tab=tab).first()
        for ls in licenses:
            License.objects.create(name=ls.model, description=ls.description, sheet=sheet_model)

    def find_sv(code):
        sv = Product.objects.filter(code=code).first()
        return sv

    if tab.controller == "LG BECON CONTROLLER":
        licenses = []
        sv = Product.objects.filter(code="SV-LG").first()

        Points.objects.create(name = sv.model,sheet = sheet_model)
        # print_title(sv.model,"B3",2)
        print_column(sheet,[sv.model,sv.description],3,2,config_style_titles2)
        

        if total_points_divice <= 2500:
            l = Product.objects.filter(code="LBML-LG").first()
            licenses.append(l)
        elif total_points_divice > 2500 and total_points_divice <= 100000:
            l = Product.objects.filter(code="LBMN-LG").first()
            licenses.append(l)
        
        calc_trafo(sheet,2,6,[sv],tab,project)
        print_row(sheet,[ls.model for ls in licenses],6,2)
        print_row(sheet,[ls.description for ls in licenses],6,3)
        save_licenses(licenses)
        



    elif tab.controller == "JCI FACILITY EXPLORER":
        licenses = []
        sv = Product.objects.filter(model="FX80").first()
        # sheet_model = Sheet.objects.create(name= sheet.title, project = tab.project )
        Points.objects.create(name = sv.model,sheet = sheet_model)
        lb =  Product.objects.filter(code="LB-FE").first()
        # print_title(sv.model,"B3",2)
        print_column(sheet,[sv.model,sv.description],3,2,config_style_titles2)
        count_devices = c_quantity
        licenses.append(lb)
        if count_devices > 0 and count_devices <=5 and total_points_divice > 0 and total_points_divice <= 250:
            lc5 =Product.objects.filter(code="LC5-FE").first() 
            licenses.append(lc5)
        elif count_devices <=10 and total_points_divice <= 500:
            lc10 =Product.objects.filter(code="LC10-FE").first() 
            licenses.append(lc10)
        elif count_devices <=25 and total_points_divice <= 1250:
            lc25 =Product.objects.filter(code="LC25-FE").first() 
            licenses.append(lc25)
        elif count_devices <=100 and total_points_divice <= 5000:
            lc100 =Product.objects.filter(code="LC100-FE").first() 
            licenses.append(lc100)

        if count_devices <=5 or count_devices <= 9:
            lm0509 = Product.objects.filter(code="LM010509-FE").first()
            licenses.append(lm0509)
        elif count_devices <=10 or count_devices <= 24:
            lm1024 = Product.objects.filter(code="LM011024-FE").first()
            licenses.append(lm1024)
        elif count_devices <=25 or count_devices <= 99:
            lm2599 = Product.objects.filter(code="LM012599-FE").first()
            licenses.append(lm2599)
        print_row(sheet,[ls.model for ls in licenses],6,2)
        print_row(sheet,[ls.description for ls in licenses],6,3)
        calc_trafo(sheet,2,6,[sv],tab,project)
        save_licenses(licenses)

    elif tab.controller == "JCI METASYS":
        sv=None
        total_divices = c_quantity + total_points_divice

        if total_divices > 0 and total_divices <=60:
            sv = find_sv("SV50-M")
        elif total_divices > 60 and total_divices <= 110:
            sv = find_sv("SV10-M")
        elif total_divices > 110 and total_divices <= 150:
            sv = find_sv("SV100-M")
        elif total_divices > 150 and total_divices <= 200:
            sv = find_sv("SV200-M")  
       

        if not sv == None:
            # sheet_model = Sheet.objects.create(name= sheet.title, project = tab.project )
            Points.objects.create(name = sv.model,sheet = sheet_model)

            calc_trafo(sheet,2,6,[sv],tab,project)
            # print_title(sv.model,"C3",2)
            print_column(sheet,[sv.model,sv.description],3,2,config_style_titles2)
        
    
def clear_cell(sheet,columns):

    

    for column in columns:
        cells = sheet[column]
        for cell in cells:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

def modify_point_file(project,sheet):

    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    config_style_titles2 = {
                            'font':False,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': True,
                            'h':"center",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    
    sheets = project.sheet.all()
    columns = ["K","L","M","N","O","P","Q","R"]
    if not sheet.title == "Supervisor":
        clear_cell(sheet,columns)
    else:
        columns = ["B","C"]
        clear_cell(sheet,columns)



    points = Points.objects.filter(sheet__in = sheets) 
   

    sh_model = None
    
    for sh in sheets:
        if sh.name == sheet.title:
            sh_model = sh
            break
    

    sh_points = [ point for point in points if sh_model.id == point.sheet.id ]

    total_points = [ point for point in sh_points if point.name == "TOTAL_POINTS" ]
     
    new_total_points = None
    if len(total_points) != 0:
        new_total_points = [total_points[0].eu,total_points[0].ed,total_points[0].sa,total_points[0].sd,total_points[0].sc]


    points_cllers_exp = [ point for point in sh_points if point.name != "TOTAL_POINTS" ]
    
    dict_points_cllers_exp = [{
            'is_controller':module.is_controller,
            'model':module.name,
            'points':[module.eu,module.ed,module.sa,module.sd,module.sc]
        }
        for module in points_cllers_exp
        ]
    sort_dict_points_cllers_exp = []

    for dict in dict_points_cllers_exp:
        if dict["is_controller"]:
            sort_dict_points_cllers_exp.insert(0,dict)
        else:
            sort_dict_points_cllers_exp.append(dict)
    

    if new_total_points:
        print_points_cller_exp_trf(sheet,new_total_points,sort_dict_points_cllers_exp,sh_model,project,True)

    if sheet.title == "Supervisor":
        all_licenses = License.objects.all()
        
        licenses = [ license for license in all_licenses if license.sheet.id == sh_model.id ]
        
        svs = [ item for item in sh_points if item.sheet.id == sh_model.id ]
          
        row = 4 + len(svs) 

        print_column(sheet,["SUPERVISOR","DESCRIPCIÓN"],2,2,config_style_titles1)
        print_column(sheet,["LICENCIA","DESCRIPCIÓN"],row,2,config_style_titles1)

        print_row(sheet,[ sv.name for sv in svs],3,2,config_style_titles2)
        print_row(sheet,[ (Product.objects.filter(model=sv.name).first()).description for sv in svs],3,3,config_style_titles2)
        
        print_row(sheet,[ license.name for license in licenses],row +1,2)
        print_row(sheet,[ license.description for license in licenses],row+1,3)
             
def style_cells(cells, config=None):
    if config is None:
            config = {
                'font': False,
                'bold': False,
                'font_color': "000000",
                'border': False,
                'alig': False,
                'wrap': False,
                'h':"left",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    for cell in cells:
        if config.get('font', False):
                cell.font = Font(bold=config.get('bold',False), color=config.get('font_color', "000000"))

        if config.get('border', False):
            cell.border = thin_border

        if config.get('alig', False):
            cell.alignment = Alignment(wrap_text=config.get('wrap',False),horizontal=config.get('h',"left"), vertical=config.get('v',"center"))


        if config.get('fill', False):
            cell.fill = PatternFill(start_color=config.get('fill_color', "FFFFFF"), end_color=config.get('fill_color', "FFFFFF"), fill_type="solid")



#funcion que guarda los item generado en la oferta
def save_offer_product(tab_obj, project, product, measure, quantity, sale_price_usd, total_value, section, porcent=0):

    if product and project or tab_obj: 
        generateoffer = GeneratedOffer.objects.filter(product=product, tab=tab_obj, project= project).first()
    else:
        generateoffer = GeneratedOffer.objects.filter(project = project, section=section, measure = measure).first()


    if generateoffer:
        generateoffer.product = product
        generateoffer.measure = measure
        generateoffer.quantity = quantity
        generateoffer.unit_value = sale_price_usd
        generateoffer.total_value = total_value
        generateoffer.tab = tab_obj
        generateoffer.project = project
        generateoffer.section = section
        generateoffer.porcent = porcent
        generateoffer.save()
    else:
        GeneratedOffer.objects.create(product=product,
                                        measure = measure,
                                        quantity=quantity,
                                        unit_value = sale_price_usd,
                                        total_value = total_value,
                                        tab= tab_obj,
                                        project = project,
                                        section = section,
                                        porcent = porcent
                                        )
        
def generate_offer_(project):

    sheets = Sheet.objects.filter(project = project).all()
    modules = Points.objects.filter(sheet__in = sheets)
    tabs = Tabs.objects.filter(project = project).all()
    divices = Divice.objects.filter(tab__in = tabs).all()
    licenses = License.objects.filter(sheet__in = sheets).all()
     
    def clean_modules(modules):# quita la fila con "TOTAL_POINTS"
        return [ module for module in modules if not module.name == "TOTAL_POINTS"]
    

    modules_cleaned = clean_modules(modules) # lista de modulos sin total_points



    modules_cleaned += licenses #Agrega las licencias a los modulos controaldores/expansiones/supervisores

    

    def count_elements(data): # cuenta los elementos
        return Counter(data)

    modules_per_tab = [{ # organiza los modulos controaldores/expansiones/supervisores por tablero
        "tab":sheet.tab.tab_name,
        "related_items":set([module.name for module in modules_cleaned if sheet.id == module.sheet.id]),
        "counter": count_elements([module.name for module in modules_cleaned if sheet.id == module.sheet.id])
    }
    for sheet in sheets
    ]


    divices_per_tab = [ # organiza los instrumentos sensores actuadores por tablero
        {"tab":tab.tab_name,
         "related_divices":set([ divice.model for divice in divices if tab.id == divice.tab.id ]),
         "counter":count_elements([ divice.model for divice in divices if tab.id == divice.tab.id ])
         }
     for tab in tabs 
    ]



    measure = "UND"
    len_item = 0

    subtotal_usd_modules = 0
    subtotal_usd_inst = 0
    subtotal_usd_ipp = 0


    # genera los controladores/expansiones/supervisores
    for dict in modules_per_tab:
        count = 0
        tab_obj = Tabs.objects.filter(tab_name = dict.get('tab'),project=project ).first()
    
        for item in dict['related_items']:
            
            product = Product.objects.filter(model=item).first()

            quantity = dict["counter"][f"{item}"]

            sale_price_usd =  product.sale_price 

            total_value = sale_price_usd * quantity

            subtotal_usd_modules += total_value 

            if tab_obj:
                save_offer_product(tab_obj,project,product,measure,quantity,sale_price_usd,total_value,0)# guarda los datos generados en la base de datos
    
    subtotal_modules = GeneratedOffer.objects.filter(
                                  measure = "SUBTOTAL CONTROLES/EXPANSIONES/SUPERVISOR",
                                  project = project
                                                              ).first()
    if subtotal_modules:

        subtotal_modules.total_value = subtotal_usd_modules
        subtotal_modules.save()
    
    else:
        GeneratedOffer.objects.create(
                                  measure = "SUBTOTAL CONTROLES/EXPANSIONES/SUPERVISOR",
                                  total_value = subtotal_usd_modules,
                                  project = project,
                                  section = 10
                                  )
    
    # genera la los dispositivos sensores/ actuadores
    for dict in divices_per_tab:
     
        tab_obj = Tabs.objects.filter(tab_name = dict.get('tab'),project=project ).first()

        for item in dict['related_divices']:
          
            product = Product.objects.filter(model=item).first()  
        
            quantity = dict["counter"][f"{item}"]
               
            sale_price_usd =  product.sale_price

            total_value = sale_price_usd * quantity

            subtotal_usd_inst += total_value 
                        
            if tab_obj:
                save_offer_product(tab_obj,project,product,measure,quantity,sale_price_usd,total_value,1)# guarda los datos generados en la base de datos

    subtotal_inst = GeneratedOffer.objects.filter(
                                  measure = "SUBTOTAL INSTRUMENTOS/COFRES",
                                  project = project,
                                  ).first()
    
    if subtotal_inst:

        subtotal_inst.total_value = subtotal_usd_inst
        subtotal_inst.save()
    
    else:
        GeneratedOffer.objects.create(
                                  measure = "SUBTOTAL INSTRUMENTOS/COFRES",
                                  total_value = subtotal_usd_inst,
                                  project = project,
                                  section = 20
                                  )

    engi_pro = Product.objects.filter(code = "IP").first()
    startup = Product.objects.filter(code="PO").first()
    
    engi_pro_startup = [engi_pro,startup]

    subtotal_usd_offer = subtotal_usd_modules + subtotal_usd_inst

    engi_pro_price_usd = subtotal_usd_offer * 0.1
    
    startup_price_usd = subtotal_usd_offer * 0.15
    

    for item in engi_pro_startup:
   
        product = Product.objects.filter(model=item.model).first()  
               
        quantity = 1

        if item.code == "IP":
            sale_price_usd = round(engi_pro_price_usd,2)
        else:
            sale_price_usd = round(startup_price_usd,2)


        total_value = sale_price_usd * quantity
        subtotal_usd_ipp += total_value
        
        
        save_offer_product(None,project,product,measure,quantity,sale_price_usd,total_value,2)# guarda los datos generados en la base de datos

    subtotal_ipp = GeneratedOffer.objects.filter(
                                  measure = "SUBTOTAL INGENIERIA/PROGRAMACIÓN/PUESTA EN OPERACIÓN",
                                  project = project,
                                  ).first()
    
    if subtotal_ipp:

        subtotal_ipp.total_value = subtotal_usd_ipp
        subtotal_ipp.save()
    
    else:
        GeneratedOffer.objects.create(
                                  measure = "SUBTOTAL INGENIERIA/PROGRAMACIÓN/PUESTA EN OPERACIÓN",
                                  total_value = subtotal_usd_ipp,
                                  project = project,
                                  section = 30
                                  )
    
    subtotal = subtotal_usd_modules+subtotal_usd_inst+subtotal_usd_ipp

    discount = round(subtotal * 0,2)
    subtotal_d = subtotal - discount 
    admi = round(subtotal_d*0.07,2)
    unexpected = round(subtotal_d*0.03,2)
    utility = round(subtotal_d*0.08,2)
    new_subtotal = subtotal_d + admi + unexpected + utility 
    iva = (new_subtotal * 0.19)
    total_value =  round(new_subtotal + iva,2)

    subtotals = [{
        'title':"SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",
        'value':subtotal,
        'porcen':None
    },
    {
        'title':"DESCUENTO",
        'value':discount,
        'porcen':"0%"
    },
    {
        'title':"SUBTOTAL MENOS (-) DESCUENTO",
        'value':subtotal_d,
        'porcen':None
    },
    {
        'title':"ADMINISTRACIÓN",
        'value':admi,
        'porcen':"7%"
    },

     {
        'title':"IMPREVISTROS",
        'value':unexpected,
        'porcen':"3%"
    },

     {
        'title':"UTILIDADES",
        'value':utility,
        'porcen':"8%"
    },
     {
        'title':"SUBTOTAL COSTO DIRECTO + INDIRECTO",
        'value':new_subtotal,
        'porcen':None
    },
    {
        'title':"IVA/UTILIDAD",
        'value':iva,
        'porcen':"19%"
    },

    {
        'title':"VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",
        'value':total_value,
        'porcen':None
    },

    ]

    def convert_to_integer(string):
        total = ''
        if string != None :
            for s in string:
                if s.isdigit():
                    total += s

            total = int(total)        
        else:
            total = 0 
            
            
        return total

    def print_subtotals(array):
        for item in array:
            save_offer_product(None,project,None,item['title'],0,0,item['value'],3,convert_to_integer(item['porcen']))
                     
    print_subtotals(subtotals)




def print_offer(sheet,project):

    sheets = Sheet.objects.filter(project = project).all()
    # modules = Points.objects.filter(sheet__in = sheets)
    tabs = Tabs.objects.filter(project = project).all()
    items_offer = GeneratedOffer.objects.filter(project = project, tab__isnull = False).all()
    engi_pro_startup = GeneratedOffer.objects.filter(section = 2,project = project).all()
    # divices = Divice.objects.filter(tab__in = tabs).all()
    # licenses = License.objects.filter(sheet__in = sheets).all()
    # def clean_modules(modules):# quita el fila con "TOTAL_POINTS"
    #     return [ module for module in modules if not module.name == "TOTAL_POINTS"]
    
    # modules_cleaned = clean_modules(modules)
    
    # modules_cleaned += licenses



    cell_dimension = ["C","D","E","F","G","H","I"]

    for cell in cell_dimension:
        sheet.column_dimensions[f'{cell}'].width = 20

    titles = ["ITEM",
              "MODELO",
              "DESCRIPCIÓN",
              "UNIDAD",
              "CANTIDAD",
              "VALOR UNITARIO (USD)",
              "VALOR TOTAL (USD)",
            ]
    
    company_info = [
        "TECHNOLOGY AND SYSTEMS INDUSTRIES SAS ",
        "NIT. 901078983",
        "CARRERA 46 165 19",
        "TEL 3230995-3163267738"
    ]
    
    client_info_titles = ["OFERTA COMERCIAL",
                   "Cliente",
                   "NIT",
                   "Fecha"
                   ]
   
    client_info_values = [
                    project.code,
                    # None,
                   project.company_name,
                   project.nit,
                   f"{date.today()}",
                   ]
    
    config_style_titles1 = {
                            'font':True,
                            'bold': True,
                            'font_color':"FFFFFF",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"538DD5"
                        }
    
    config_style_titles2 = {
                            'font':True,
                            'bold': True,
                            'font_color':"FFFFFF",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"808080"
                        }
    
    config_style_titles3 = {
                            'font':True,
                            'bold': True,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': False,
                            'h':"center",
                            'v':"center",
                            'fill':True,
                            'fill_color':"DAD7D7"
                        }
    
    config_style_titles4 = {
                            'font':False,
                            'bold': False,
                            'font_color':"000000",
                            'border':True,
                            'alig':True,
                            'wrap': True,
                            'h':"center",
                            'v':"center",
                            'fill':False,
                            'fill_color':"DAD7D7"
                        }
    
    

    print_row(sheet,client_info_titles,3,8,config_style_titles3)
    print_row(sheet,client_info_values,3,9,config_style_titles4)
    

    print_row(sheet,company_info,3,5,{
                'font': True,
                'bold': True,
                'font_color': "000000",
                'border': False,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "DAD7D7",
            })
    
    print_column(sheet,["CONTROLADORES / EXPANSIONES / SUPERVISORES"],8,3,config_style_titles1)
    sheet.merge_cells('C8:I8')
    sheet.row_dimensions[8].height = 25

    def count_elements(data):
        return Counter(data)
    
    # modules_per_tab = []


    # for tab in tabs:
    #     tab_dict = {}
    #     tab_dict['tab'] = tab
    #     items = []
    #     for item in items_offer:
    #         if tab.id == item.tab.id:
    #             if item.section == 0:
    #                 items.append(item)
    #     tab_dict['related_items'] = items
    #     modules_per_tab.append(tab_dict)


    modules_per_tab = [{
        "tab":tab,
        "related_items":[item for item in items_offer if tab.id == item.tab.id and item.section == 0 ],
        # "counter": count_elements([item for item in items_offer if tab.id == item.tab.id])
    }
    for tab in tabs
    ]
    
    divices_per_tab = [
        {"tab":tab,
        "related_items":[item for item in items_offer if tab.id == item.tab.id and item.section == 1 ],
        #  "counter":count_elements([ divice.model for divice in divices if tab.id == divice.tab.id ])
         }
     for tab in tabs 
    ]


    sheet.column_dimensions["E"].width = 60
    sheet.column_dimensions["C"].width = 5
    
    measure = "UND"
    len_item = 0
    for dict in modules_per_tab:
        count = 0
        cell = sheet.cell(row=9+len_item, column=3, value=(dict['tab']).tab_name) #imprime el tablero
        # tab_obj = Tabs.objects.filter(tab_name = dict.get('tab'),project=project ).first()
        sheet.merge_cells(f"C{9+len_item}:I{9+len_item}")
        style_cells([cell],config_style_titles2)#le da estilo a la celda de tab
        print_column(sheet,titles,10+len_item,3,config_style_titles3)#imprime los titulos
        # for i,item in enumerate(item['related_items'], start=11+len_item):
        for i,item in enumerate(dict['related_items'], start=11+len_item):
         
            count+=1
            cell_item = sheet.cell(row=i, column=3, value=count)#imprime el consecutivo
            style_cells([cell_item],config_style_titles3)
            cell_model = sheet.cell(row=i, column=4, value=item.product.model)#imprime el modelo del producto
        #     # product = Product.objects.filter(model=item).first()
            cell_description = sheet.cell(row=i, column=5, value=item.product.description)#imprime la descripción del producto
        
            cell_und = sheet.cell(row=i, column=6, value="UND")# imprime la unidad de medida
            
            # quantity = item.quantity

            cell_quantity = sheet.cell(row=i, column=7, value=item.quantity)# imprime la cantidad
            
            # sale_price_usd =  item.unit_value

            cell_price_usd = sheet.cell(row=i, column=8, value=item.unit_value)#imprime el valor del producto
            
            # total_value = sale_price_usd * quantity

            cell_total_value_usd = sheet.cell(row=i, column=9, value=item.total_value)#imprime el valor total de los producto
            
            style_cells([cell_model,
                        cell_description,
                        cell_und,
                        cell_quantity,
                        cell_price_usd,
                        cell_total_value_usd,
                        ],config_style_titles4)# le da estilos a las celdas
            
        len_item += len(dict['related_items']) + 2     
     
    subtotal_usd_modules = count_value_row(sheet,9,11,9 + len_item)

    print_column(sheet,[None,
                        None,
                        "SUBTOTAL ( USD)",
                        None,
                        None,
                        None,
                        subtotal_usd_modules],9 + len_item,3,config_style_titles3)
    
    len_item_two = len_item + 3

    sheet.row_dimensions[9 + len_item_two-1].height = 25

    print_column(sheet,["INTRUEMENTOS / COFRES"],9 + len_item_two-1,3,config_style_titles1)
    sheet.merge_cells(f"C{9 + len_item_two-1}:I{9 + len_item_two-1}")

    for dict in divices_per_tab:
        count = 0
        cell_tab = sheet.cell(row=9+len_item_two, column=3, value=(dict.get('tab')).tab_name)
        sheet.merge_cells(f"C{9+len_item_two}:I{9+len_item_two}")
        # tab_obj = Tabs.objects.filter(tab_name = dict.get('tab'),project=project ).first()
        style_cells([cell_tab],config_style_titles2)
        print_column(sheet,titles,9+len_item_two+1,3,config_style_titles3)
        for i,item in enumerate(dict['related_items'], start=11+len_item_two):
            count+=1
         
            cell_item = sheet.cell(row=i, column=3, value=count)

            style_cells([cell_item],config_style_titles3)

            # product = Product.objects.filter(model=item).first()  
            # divice = Divice.objects.filter(model=item).first() 
            # if not (item == "AI" or item == "BI" or item == "AO" or item == "BO"):
            
            cell_model = sheet.cell(row=i, column=4, value=item.product.model) # imprime el modelo del dispositivo
            
            cell_description = sheet.cell(row=i, column=5, value=item.product.product_name) # Imprime la descripción
           
            cell_und = sheet.cell(row=i, column=6, value="UND") # imprime la unidad de medida
           
            # quantity = dict["counter"][f"{item}"]

            # quantity = item.quantity # cantidad del producto

            cell_quantity = sheet.cell(row=i, column=7, value=item.quantity) # imprime la cantidad
               
            # sale_price_usd =  item.product.sale_price # precio del producto

            cell_price_usd = sheet.cell(row=i, column=8, value=item.unit_value)

            # total_value = sale_price_usd * quantity # precio total 
            
            cell_total_value_usd = sheet.cell(row=i, column=9, value=item.total_value) # imprime el precio total
            
            style_cells([cell_model,
                         cell_description,
                         cell_und,
                         cell_quantity,
                         cell_price_usd,
                         cell_total_value_usd,
                         ],config_style_titles4)
            
        len_item_two += len(dict.get('related_items')) + 2       
    
    
    subtotal_usd_inst = count_value_row(sheet,9,11+len_item,9 +  len_item_two)


    print_column(sheet,[None,
                        None,
                        "SUBTOTAL ( USD )",
                        None,
                        None,
                        None,
                        subtotal_usd_inst],9 +  len_item_two,3,config_style_titles3)
    
    print_column(sheet,
                 ["INGENIERÍA / PROGRAMACIÓN / PUESTA EN OPERACIÓN"],
                 11 +  len_item_two,3,
                 config_style_titles1)
    
    sheet.row_dimensions[11 +  len_item_two].height = 25
    
    sheet.merge_cells(f"C{11+len_item_two}:I{11+len_item_two}")
    
    # engi_pro = Product.objects.filter(code = "IP").first()
    # startup = Product.objects.filter(code="PO").first()
    
    print_column(sheet,titles,12+len_item_two,3,config_style_titles3)
    
    # engi_pro_startup = [engi_pro,startup]
    
    count = 0

    # # subtotal_usd_offer = subtotal_usd_modules + subtotal_usd_inst

    # # engi_pro_price_usd = subtotal_usd_offer * 0.1
    
    # # startup_price_usd = subtotal_usd_offer * 0.15
    
    for i,item in enumerate(engi_pro_startup,start=13+len_item_two):
        count+=1
        cell_item = sheet.cell(row=i, column=3, value=count)

        style_cells([cell_item],config_style_titles3)

        cell_model = sheet.cell(row=i, column=4, value=item.product.model)

        # product = Product.objects.filter(model=item.model).first()  
       
        cell_description = sheet.cell(row=i, column=5, value=item.product.description)
        
        cell_und = sheet.cell(row=i, column=6, value="UND")


        cell_quantity = sheet.cell(row=i, column=7, value=item.quantity)
        # if item.code == "IP":
        #     sale_price_usd = round(engi_pro_price_usd,2)
        # else:
        #     sale_price_usd = round(startup_price_usd,2)

        cell_price_usd = sheet.cell(row=i, column=8, value=item.unit_value)
        
        cell_total_value_usd = sheet.cell(row=i, column=9, value= item.total_value)
        
    
        style_cells([cell_model,
                        cell_description,
                        cell_und,
                        cell_quantity,
                        cell_price_usd,
                        cell_total_value_usd,
                        ],config_style_titles4)
    
    subtotal_usd_ipp = count_value_row(sheet,9,13+len_item_two, 13+len_item_two + count)


    print_column(sheet,[
        None,
        None,
        "SUBTOTAL ( USD)",
        None,
        None,
        None,
        subtotal_usd_ipp],
        13 + len_item_two + count,3,config_style_titles3)
    
    # subtotal = subtotal_usd_modules+subtotal_usd_inst+subtotal_usd_ipp

    # # discount = round(subtotal * 0,2)
    # # subtotal_d = subtotal - discount 
    # # admi = round(subtotal_d*0.07,2)
    # # unexpected = round(subtotal_d*0.03,2)
    # # utility = round(subtotal_d*0.08,2)
    # # new_subtotal = subtotal_d + admi + unexpected + utility 
    # # iva = (new_subtotal * 0.19)
    # # total_value =  round(new_subtotal + iva,2)

    subtotal = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",project= project).first()
    descuento = GeneratedOffer.objects.filter(measure__icontains="DESCUENTO",project= project).first()
    subtotal_descuento = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL MENOS (-) DESCUENTO",project= project).first()
    admin = GeneratedOffer.objects.filter(measure__icontains="ADMINISTRACIÓN",project = project).first()
    imprevistos = GeneratedOffer.objects.filter(measure__icontains="IMPREVISTROS",project = project).first()
    utilidades = GeneratedOffer.objects.filter(measure__icontains="UTILIDADES",project = project).first()
    subtotal_directo_indirecto = GeneratedOffer.objects.filter(measure__icontains="SUBTOTAL COSTO DIRECTO + INDIRECTO",project = project).first()
    iva_utilidad = GeneratedOffer.objects.filter(measure__icontains="IVA/UTILIDAD",project = project).first()
    valor_total = GeneratedOffer.objects.filter(measure__icontains="VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",project = project).first()


    subtotals = [{
        'title':"SUBTOTAL SUMINISTROS Y DESARROLLO ( USD )",
        'value':subtotal.total_value,
        'porcen':None
    },
    {
        'title':"DESCUENTO",
        'value':descuento.total_value,
        'porcen':f"{descuento.porcent}%"
    },
    {
        'title':"SUBTOTAL MENOS (-) DESCUENTO",
        'value':subtotal_descuento.total_value,
        'porcen':None
    },
    {
        'title':"ADMINISTRACIÓN",
        'value':admin.total_value,
        'porcen':f"{admin.porcent}%"
    },

     {
        'title':"IMPREVISTROS",
        'value':imprevistos.total_value,
        'porcen':f"{imprevistos.porcent}%"
    },

     {
        'title':"UTILIDADES",
        'value':utilidades.total_value,
        'porcen':f"{utilidades.porcent}%"
    },
     {
        'title':"SUBTOTAL COSTO DIRECTO + INDIRECTO",
        'value':subtotal_directo_indirecto.total_value,
        'porcen':None
    },
    {
        'title':"IVA/UTILIDAD",
        'value':iva_utilidad.total_value,
        'porcen':f"{iva_utilidad.porcent}%"
    },

    {
        'title':"VALOR TOTAL (USD) DÓLAR ESTADOUNIDENSE",
        'value':valor_total.total_value,
        'porcen':None
    },

    ]

    # # def convert_to_integer(string):
    # #     total = ''
    # #     if string != None :
    # #         for s in string:
    # #             if s.isdigit():
    # #                 total += s

    # #         total = int(total)        
    # #     else:
    # #         total = 0 
            
    # #     return total

    
    def print_subtotals(sheet,array):
        for i,item in enumerate(array,start=15):
                  
            print_column(sheet,[None,
                                None,
                                item['title'],
                                None,
                                None,
                                item['porcen'],
                                item['value']],
                                i + len_item_two + count,
                                3,
                                config_style_titles3)
            
    print_subtotals(sheet,subtotals)


def print_notes(sheet,project):


    sheet.column_dimensions['C'].width = 120
    # sheet.row_dimensions[6].height = 200
    # sheet.row_dimensions[10].height = 200
    alcances = Note.objects.filter(tag="ALCANCES").first()
    no_incluye = Note.objects.filter(tag="NO INCLUYE").first()
    condiciones_comerciales = Note.objects.filter(tag="CONDICIONES COMERCIALES").first()
    notas_aclaraciones = Note.objects.filter(tag="NOTAS Y ACLARACIONES").first()

    config_style_titles1 = {
                'font': True,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': False,
                'wrap': True,
                'h':"left",
                'v':"center", 
                'fill': True,
                'fill_color': "DAD7D7",
            }
    config_style_titles2 = {
                'font': False,
                'bold': False,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"left",
                'v':"top", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    
    print_row(sheet,["ALCANCES"],5,3,config_style_titles1)
    if alcances:
        print_row(sheet,[alcances.description],6,3,config_style_titles2)
    
    sheet.merge_cells("C6:C20")

    print_row(sheet,["NO INCLUYE"],23,3,config_style_titles1)
    if no_incluye:
        print_row(sheet,[no_incluye.description],24,3,config_style_titles2)
    sheet.merge_cells("C24:C60")

    print_row(sheet,["CONDICIONES COMERCIALES"],63,3,config_style_titles1)
    if condiciones_comerciales:
        print_row(sheet,[condiciones_comerciales.description],64,3,config_style_titles2)
    sheet.merge_cells("C64:C85")


    print_row(sheet,["NOTAS Y ACLARACIONES"],88,3,config_style_titles1)
    if notas_aclaraciones:
        print_row(sheet,[notas_aclaraciones.description],89,3,config_style_titles2)
    sheet.merge_cells("C89:C110")

def merge_col(sheet,col1,col2,start,end):
    for i in range(start,end + 1):
        sheet.merge_cells(f"{col1}{i}:{col2}{i}")
    
def print_remission(sheet,remission):
    config_style_titles1 = {
                'font': True,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    config_style_titles2 = {
                'font': False,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    config_style_titles3 = {
                'font': False,
                'bold': True,
                'font_color': "000000",
                'border': False,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    print_row(sheet,["REMISIÓN DE MATERIALES"],1,5,config_style_titles1)
    sheet.merge_cells("E1:H3")
    sheet.merge_cells("B1:D3")

    print_row(sheet,["CODIGO","VERSIÓN","FECHA"],1,9,config_style_titles1)

    print_row(sheet,["SGC-FO-004","VERSIÓN 02","01/11/2018"],1,10,config_style_titles1)
    sheet.merge_cells("E1:H3")
    title1 =[
        "Ciudad",
        "Razon Social",
        "NIT",
        "Lugar de entrega del material"
    ]

    title1_values =[
        remission.city,
        remission.company,
        remission.nit,
        remission.location
    ]

    title2 =[
        "Remision No:",
        "Proveedor:",
        "Fecha",
        "Pedido No:",
        "Obra:",
        "Ingeniero responsable de la obra:"
    ]
    title2_values =[
        remission.number,
        remission.company,
        remission.date.strftime("%d-%m-%Y"),
        remission.order_number,
        remission.contruction_site,
        remission.responsible
    ]

    title3 =[
        "Cantidad Entregado",
        "Referencia",
        None,
        "Descripción",
        None,
        None,
        None,
        None,
        "Observación",
    ]
    print_row(sheet,title1,7,2,config_style_titles1)
    
    merge_col(sheet,"B","D",7,10)    
    print_row(sheet,title1_values,7,5,config_style_titles1)
    merge_col(sheet,"E","H",7,10) 
    print_row(sheet,title2,5,9,config_style_titles1)
    print_row(sheet,title2_values,5,10,config_style_titles2)
    
    print_column(sheet,title3,13,2,config_style_titles1)
    sheet.merge_cells("C13:D13")
    sheet.merge_cells("E13:I13")
    
    remission_product = ProductSent.objects.filter(remission = remission).all()
    
    for  i,product in enumerate(remission_product,start=14):
        index_cell = sheet.cell(row=i, column=2, value=product.quantity)
        cell_model = sheet.cell(row=i, column=3, value=product.product.model)
        cell_description = sheet.cell(row=i, column=5, value=product.product.description)
        cell_observation = sheet.cell(row=i, column=10, value=None)
        style_cells([index_cell],config_style_titles1)
        style_cells([cell_model,cell_description,cell_observation],config_style_titles2)      
        sheet.merge_cells(f"E{i}:I{i}")
        sheet.merge_cells(f"C{i}:D{i}")
    
    row1 = 13 + len(remission_product) + 4
    sheet.merge_cells(f"C{row1}:D{row1}")
    print_column(sheet,["Entrega","________________"],row1,2,config_style_titles3)
    
    print_column(sheet,["Aceptado","________________"],row1,9,config_style_titles3)
    
    sheet.column_dimensions['D'].width = 10
    path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'img','logo2.png')

    logo = Image(path)
    sheet.add_image(logo,"B1")


def row_dimen_height(sheet,tam,start,end):
        for i in range(start,end + 1):
            sheet.row_dimensions[i].height = tam

def print_order(sheet,order):
    config_style_titles1 = {
                'font': True,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "FFFFFF",
            }
    config_style_titles2 = {
                'font': True,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': True,
                'fill_color': "F1F1F1",
            }
    config_style_titles3 = {
                'font': False,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"center",
                'v':"center", 
                'fill': False,
                'fill_color': "DAD7D7",
            }
    config_style_titles3 = {
                'font': False,
                'bold': True,
                'font_color': "000000",
                'border': True,
                'alig': True,
                'wrap': True,
                'h':"left",
                'v':"center", 
                'fill': False,
                'fill_color': "DAD7D7",
            }
    # sheet.column_dimensions['J'].width = 20
    sheet.row_dimensions[1].height = 50
    print_column(sheet,["ORDEN DE COMPRA Y/O SERVICIOS"],1,6,config_style_titles1)
    sheet.merge_cells("F1:K1")
    sheet.merge_cells("A1:E1")
    title1 = [
        "PROVEEDOR / CONTRATISTA",
        "NIT/C.C.",
        "DIRECCIÓN",
        "TELEFONO",
        "CLIENTE"
    ]
    title2 = [
        "ORDEN DE COMPRA No.",
        "FECHA",
        "CENTRO DE COSTO",
        "INTERVENTOR",
        "SUPERVISOR"
    ]

    title3 = [
        "ITEM",
        "UNIDAD",
        "DESCRIPCION",
        None,
        None,
        None,
        None,
        "CANTIDAD",
        None,
        "V.UNITARIO",
        None,
        "V.TOTAL",
        None
    ]
    title1_values = [
        order.supplier,
        order.nit,
        order.address,
        order.phone,
        order.customer
    ]
    title2_values = [
        order.code,
        order.date.strftime("%d-%m-%Y"),
        order.cost_center,
        order.inspector,
        order.supervisor
    ]
    print_row(sheet,title1,3,1,config_style_titles2)
    merge_col(sheet,"A","E",2,8) 
    print_row(sheet,title1_values,3,6,config_style_titles3)
    merge_col(sheet,"F","H",3,7) 
    print_row(sheet,title2,3,10,config_style_titles2)
    merge_col(sheet,"J","K",3,7) 
    print_row(sheet,title2_values,3,12,config_style_titles3)
    merge_col(sheet,"L","M",3,7)
    print_row(sheet,["SGC-F0-011"],1,12,config_style_titles1)
    merge_col(sheet,"L","M",1,12)

    sheet.merge_cells("C9:G9")
    sheet.merge_cells("H9:I9")
    sheet.merge_cells("J9:K9")
    sheet.merge_cells("L9:M9")
    print_column(sheet,title3,9,1,config_style_titles2)



    order_product = OrderProduct.objects.filter(order = order).all()
    
    count = 0
    subtotal_value = 0
    for  i,product in enumerate(order_product,start=10):
        count += 1
        index_cell = sheet.cell(row=i, column=1, value=count)
        cell_model = sheet.cell(row=i, column=3, value=product.product.model)
        cell_quantity = sheet.cell(row=i, column=8, value=product.quantity)
        cell_price = sheet.cell(row=i, column=10, value=product.price)
        total_value = product.quantity * product.price
        cell_total_value = sheet.cell(row=i, column=12, value=total_value)
        subtotal_value += total_value
        cell_unit = sheet.cell(row=i, column=2, value="UND")
        # cell_description = sheet.cell(row=i, column=5, value=product.product.description)
        # cell_observation = sheet.cell(row=i, column=10, value=None)
        # style_cells([index_cell],config_style_titles1)
        style_cells([index_cell,cell_model,cell_quantity,cell_price,cell_total_value,cell_unit],config_style_titles1)      
        sheet.merge_cells(f"C{i}:G{i}")
        sheet.merge_cells(f"H{i}:I{i}")
        sheet.merge_cells(f"J{i}:K{i}")
        sheet.merge_cells(f"L{i}:M{i}")
    
    row1 = 9 + len(order_product) + 1

    print_column(sheet,["SUBTOTAL",None,subtotal_value],row1,10,config_style_titles2)
    sheet.merge_cells(f"J{row1}:K{row1}")
    sheet.merge_cells(f"L{row1}:M{row1}")

    conditions = Note.objects.filter(tag="CONDICIONES OC").first()
    if conditions:
        print_row(sheet,[conditions.description],row1+2,1,config_style_titles3)
    else:
        description =  """
            			
                1. Tipo de contrato:XXXXX											
                2. El valor ofertado es a todo costo, el cual cubre todo el objeto negociado y los  sobre costos que se llegaren a presentar son responsabilidad de XXXXXX											
                3. El proveedor deberá emplear toda su experiencia y core de negocio, para cumplir con el objeto de la presente orden de compra, en términos de calidad											
                oportunidad, eficiencia y servicio post-venta											
                "Nota Con la aceptacion de este documento el proveedor se obliga a cumplir con todas las normas de seguridad industrial y ocupacional establecidas en la ley para las personas para las personas que realizan trabajos en altura, como a nivel de piso, tanto para personas como para los materiales y equipos que se empleen en el desarrollo de este contrato.
                Para la ejecucion de las catividades, no se permitora el ingreso del personal ue no este afilado al sistema de seguridad social (EPS, ARL Y PENSION) y cuente con la cobertura acorde con los trabajos contratados. Igualmente para realizar los pagos se debe adjuntar copia de las planillas con los pagos realizados. "											
                4. Este contrato, no es un contrato laboral, por lo que no existe un relación entre el CONTRATANTE - T&S INC y el CONTRATISTA. Por ende, no											
                hay relación directa, ni existe responsabilidad entre las partes											
                5. FORMA DE PAGO: las partes acuerdan:											
                                                            
                6. Especificaciones, Todas las contenidas en la cotizacion XXXXXX											
                Representante legal XXXXXX											
                7. PÓLIZAS DE LA PRESENTE ORDEN DECOMPRA
        
        """
        conditions = Note.objects.create(tag="CONDICIONES OC", description = description)
        print_row(sheet,[conditions.description],row1+2,1,config_style_titles3)

    row2 = row1 + 25
    sheet.merge_cells(f"A{row1+2}:M{row2}")

    title4 = ["PÓLIZAS",
              None,
              None,
              None,
              None,
              "SI",
              "NO",
              "CONDICIONES"]

    poliza = ["Buen Manejo Anticipo",
              "Cumplimiento y tiempo de entrega",
              "Salarios y prestaciones sociales",
              "Seguro de Responsabilidad Civil Extracontractual",
              "Garantia, Calidad y Estabilidad de Obra"
              ]
    condiciones = [
        "Por valor igual al 100% del monto del anticipo y una vigencia mínima a la duración del contrato y tres meses más",
        "Por un valor asegurado equivalente al 30% del valor total del contrato IVA incluido, vuya vigencia será igual a la duración del Contrato y tres meses más",
        "Por un valor asegurado equivalente al 30% del valor total del contrato IVA incluido, vuya vigencia será igual a la duración del Contrato y tres meses más",
        "Por un valor asegurado equivalente al 30% del valor total del contrato IVA incluido, vuya vigencia será igual a la duración del Contrato y tres meses más",
        "Por un valor asegurado equivalente al 30% del valor total del contrato IVA incluido, vuya vigencia será igual a un año más tres meses más"
    ]
    row3 = row2 + 1
    print_column(sheet,title4, row3 ,1,config_style_titles2)
    row4 = row3 + 1

    print_row(sheet,poliza,row4,1,config_style_titles3)
    print_row(sheet,condiciones,row4,8,config_style_titles3)
    merge_col(sheet,"A","E",row3,row3 + len(poliza))
    merge_col(sheet,"H","M",row3,row3 + len(poliza))

    row_dimen_height(sheet,55,row4,row4 + len(poliza)-1)
    row5 = row4 + len(poliza)
    print_row(sheet,["CONDICIONES DE LA PRESENTE ORDEN DE COMPRA"],row5,1,config_style_titles2)
    sheet.merge_cells(f"A{row5}:M{row5}")

    order_conditions = ["Fecha Inicio Contrato - Orden:",
                        "Fecha Final Contrato - Orden:",
                        "Fecha de Cumplimiento y Entrega:",
                        "Lugar de Entrega:",
                        "Interventor Contrato T&S INC",
                        "Bogotá D.C.",
                        "Aplican Entregas Parciales:"
                        ]
    order_conditions_values = [None,
                               None,
                               None,
                               None,
                               None,
                               None,
                               "SI__ NO__ El proveedor podrá entregar entregas parciales"
                        ]
    row6 = row5 + 1
    print_row(sheet,order_conditions,row6,1,config_style_titles3)
    print_row(sheet,order_conditions_values,row6,7,config_style_titles3)
    merge_col(sheet,"A","F",row6,row6 + len(order_conditions))
    merge_col(sheet,"G","M",row6,row6 + len(order_conditions))

    delivery_conditions = ["1. El proveedor debe cumplir con lo solicitado en la presente orden de compra",
                           "2. Para dar cumplimiento a la forma de pago, el contratista deberá remitir adjunto a la factura, acta de entrega firmada de recibida a satisfacción por T&S INC.",
                           "3. El Proveedor deberá anexar, el respectivo pago de parafiscales del personal que empleó para el desarrollo del objeto del contrato",
                           "4. El Provedor sera responsable de suministrar todas las herramientas de trabajo requeridas para el ensamble de los tableros de control",
                           "5. El Provedor sera responsable del pago de la seguridad social tanto de el como del personal que tenga a su cargo.",
                           "6. El Provedor sera responsable de suministrar todos los elementos de seguridad (EPP) requeridos tanto por el como por el personal a su cargo para el ENSAMBLE DE LOS TABLEROS DE CONTROL"
    ]
    row7 = row6 + len(order_conditions)

    print_row(sheet,delivery_conditions,row7,7,config_style_titles3)
    merge_col(sheet,"G","M",row7,row7+len(delivery_conditions))
    row_dimen_height(sheet,55,row7,row7+len(delivery_conditions)-1)

    note = ["El incumplimiento por parte del proveedor- contratista de cualquiera de las obligaciones derivadas de la presente orden de compra, dara lugar a la imposicion de sancion penal pecunaria  a favor TECHNOLOGY AND SYSTEMS INDUSTRIES SAS por el valor de la presente orden  y la devolucion del anticipo si lo hubere y o la indemnizacion de los perjuicios que ocasione."]
    row8 = row7+len(delivery_conditions) + 2
    print_row(sheet,note,row8,1,config_style_titles3)
    merge_col(sheet,"A","M",row8,row8)
    row_dimen_height(sheet,55,row8,row8)

    signature1 = [
        "AUTORIZADO POR",
    ]
    signature2 = [
        "CONTRATISTA",
    ]

    print_column(sheet,signature1,row8+3,1,config_style_titles3)
    print_column(sheet,signature2,row8+3,8,config_style_titles3)
    merge_col(sheet,"A","F",row8+3,row8+3)
    merge_col(sheet,"H","M",row8+3,row8+3)
    row_dimen_height(sheet,30,row8+3,row8+3)

    path = os.path.join(settings.BASE_DIR, 'tsinc','static', 'img','logo2.png')

    logo = Image(path)
    sheet.add_image(logo,"A1")



    








    


    
    

 

    



        



   
        
   